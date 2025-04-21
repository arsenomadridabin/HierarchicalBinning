import re
import sys
import argparse
import json
import matplotlib.pyplot as plt
import numpy as np
import matplotlib

# Fe 5440
# Mg 6656
# Si 4608
# O 16064

"""Atom class
Fe : 1
Mg:2
Si :3
O:4
N :5
"""
ATOM_CLASS_MAPPER_NUMBER = {
        "fe" : 1,
        "mg" : 2,
        "si" : 3,
        "o" : 4,
        "n" : 5,
        "h" : 6
}



# alpha = 1000
# mg inside = 968
# mg outside = 5752

# alpha = 30
# mg inside = 942
# mg outside = 5778

#alpha =10
# mg outside = 5870
# mg inside = 850

#alpha = 1000
#fe inside = 1327
# fe outside = 4177

#alpha =30
# fe inside = 1263
#fe outside = 4241


# alpha = 10
# fe outside = 4359
# fe inside = 1145


#alpha 1000
#Si outside = 3918
# Si inside = 818


#alpha 30
# Si outside = 3955
# Si inside_count = 781

#alpha 10
# Si outside = 4036
# Si inside = 700




ATOM_CLASS_MAPPER = {
        "fe" : "Fe",
        "mg" : "Mg",
        "si" : "Si",
        "o" : "O",
        "n" : "N"
}

#Distance at which Direct Confighutation starts in XDATCAR file starts
OFFSET_DISTANCE = 7

def parsed_data(data):

    x_data = float(data[2])
    y_data = float(data[3])
    z_data = float(data[4])

    #x_cord
    if x_data > CELL_SIZE/2:
        x_data = x_data - CELL_SIZE

    #y_cord
    if y_data > CELL_SIZE/2:
        y_data = y_data - CELL_SIZE

    #z_cord
    if z_data > CELL_SIZE/2:
        z_data = z_data - CELL_SIZE



    return {
        "atom_id" : data[0],
        "atom_class" : data[1],
        "atom_coordinate" : [x_data,y_data,z_data] 
    }


def get_points(path):
    with open(path, 'r') as myfile:
        final_data = json.load(myfile)

    if type(final_data[0]) != list:
        final_data = [final_data]
    # pts = [[pt['atom_coordinate'] for pt in pts] for pts in final_data]
    pts = [[{'coordinate':pt['atom_coordinate'],'id' : pt.get('atom_id',''), 'atom_class' : pt.get('atom_class')} for pt in pts] for pts in final_data]
    return pts


def countX(lst, x):
    count = 0
    for ele in lst:
        if (ele == x):
            count = count + 1
    return count



def get_atom_distribution(data,filtering=False,atom_class=None, return_buckets=None,graph_output=None):
        import matplotlib.pyplot as plt
        my_distro = {}
        for i in range(MAX_ATOMS_PER_BUCKET+1):
            my_distro[i] = 0

        # #Only using 1 data
        # all_data = [all_data[0]]
        buckets = {}
        for i in range(NO_OF_BUCKET):
                #old way
                if not return_buckets:
                        buckets[i] = 0
                else:
                        if i in return_buckets:
                                buckets[i] = 0
        for each in data:
                if filtering:
                        val_class = each['atom_class']
                        if not val_class == atom_class:
                                continue
                try:
                        datum = each['coordinate']
                except Exception as e:
                        datum = each['atom_coordinate']
                x_cord, y_cord, z_cord = datum[0], datum[1], datum[2]
                x_offset = int((x_cord - LEFT_BOUNDARY)/BUCKET_LENGTH)
                y_offset = int((y_cord - LEFT_BOUNDARY)/BUCKET_LENGTH) 
                z_offset = int((z_cord - LEFT_BOUNDARY)/BUCKET_LENGTH)

                z_multiplier = round(NO_OF_BUCKET ** (2/3),3)
                y_multiplier = round(NO_OF_BUCKET ** (1/3),3)
                val = z_multiplier * z_offset + y_multiplier * y_offset + x_offset
                bucket_no = int((z_multiplier * z_offset + y_multiplier * y_offset + x_offset))
                if bucket_no == 8:
                        pass
                        # print("Coordinates : {} , {}, {}, {}".format(x_cord,y_cord,z_cord,val))
                try:
                    a = buckets[bucket_no]
                    a = a + 1
                    buckets[bucket_no] = a
                except Exception as e:
                        import traceback
                        # print(traceback.format_exc())
                        # print("Exception--",bucket_no)
        
        if atom_class == "1":
                file = open("buckets_fe_density.json","w")
        elif atom_class == "2":
                file = open("buckets_mg_density.json","w")
        elif atom_class == "3":
                file = open("buckets_si_density.json","w")
        elif atom_class == "4":
                file = open("buckets_o_density.json","w")
        else:
                file = open("buckets_n_density.json","w")

        json.dump(buckets,file)
        file.close()

        list_num_per_bucket = []
        for x,y in buckets.items():
            list_num_per_bucket.append(y)

        unique_elements = list(set(list_num_per_bucket))
        print("uni=",unique_elements)
        print("list_num_per_bucket=",list_num_per_bucket)
        print("my distro-",my_distro)
        for element in unique_elements:
            count = countX(list_num_per_bucket,element)
            my_distro[element] = my_distro[element] + count


        print("my_distro=",my_distro)


        final_frequency = {}

        for key,value in my_distro.items():
            final_frequency[key] = round(value,2)

        print("final-freq=",final_frequency)

        val = []
        for x,y in final_frequency.items():
            val = val + int(round(y,2)) * [x]


        final_list = []
        for i in val:
            final_list.append(int(round(i,2)))

        set_list = set(final_list)
        #old
        # length = len(set_list)

        # new
        length = max(final_list) - min(final_list)


        #Ploting
        from datetime import datetime
        now = datetime.now()
        timestamp = int(now.timestamp())
        print("final_list=",final_list)
        if 1:
                import pandas as pd
                import seaborn as sns
                try:
                        if filtering:
                                n, bins, patches = plt.hist(final_list, length, facecolor='blue', alpha=0.5,edgecolor="red",align='mid' )

                                print("atomclass={}".format(atom_class))
                                if atom_class == "1":
                                        plt.xlabel('Number of Fe atoms in a cluster')
                                        plt.ylabel('Frequency')
                                        plt.title('Distribution of Fe atoms in a cluster')
                                        plt.savefig('Subcluster_{}_{}'.format("Fe",timestamp))
                                elif atom_class == "4":
                                        plt.xlabel('Number of O atoms in a cluster')
                                        plt.ylabel('Frequency')
                                        plt.title('Distribution of 0 atoms in a cluster')
                                        plt.savefig('Subcluster_{}_{}'.format("O",timestamp))
                                elif atom_class == "2":
                                        plt.xlabel('Number of Mg atoms in a cluster')
                                        plt.ylabel('Frequency')
                                        plt.title('Distribution of Mg atoms in a cluster')
                                        plt.savefig('Subcluster_{}_{}'.format("Mg",timestamp))
                                elif atom_class == "5":
                                        plt.xlabel('Number of H atoms in a cluster')
                                        plt.ylabel('Frequency')
                                        plt.title('Distribution of H atoms in a cluster')
                                        plt.savefig('Subcluster_{}_{}'.format("H",timestamp))                                                                                           
                                else:
                                        plt.xlabel('Number of Si atoms in a cluster')
                                        plt.ylabel('Frequency')
                                        plt.title('Distribution of Si atoms in a cluster')
                                        plt.savefig('Subcluster_{}_{}'.format("Si",timestamp))
                        else:
                                print("abin")
                                # print("sssss")
                                # plt.xlabel('Number of {} atoms in a bin'.format(ATOM_CLASS_MAPPER[graph_output]))
                                # plt.ylabel('Frequency')
                                # plt.title('Distribution of {} atoms in a cell'.format(ATOM_CLASS_MAPPER[graph_output]))
                                # plt.savefig('{}_{}'.format(args.output,timestamp))
                                # print("y")
                                print("fin-->",final_list)
                                df = pd.DataFrame({
                                                                "Frequency of {} Atom".format(ATOM_CLASS_MAPPER[graph_output]) : final_list
                                                                })
                                try:
                                        plt2 = sns.histplot(data=df, x="Frequency of {} Atom".format(ATOM_CLASS_MAPPER[graph_output]), kde=True,discrete=True,fill=False).set(title='Distribution of {} Atom in a Simulation Box'.format(ATOM_CLASS_MAPPER[graph_output]))
                                        fig2 = plt2[0].get_figure()
                                        fig2.savefig("out_{}.png".format(ATOM_CLASS_MAPPER[graph_output]))
                                except Exception as e:
                                        import traceback
                                        traceback.format_exc()



                except Exception as e:
                        import traceback
                        print(traceback.format_exc())

        else:
                import pandas as pd
                import seaborn as sns
                df = pd.DataFrame({
                        "Frequency of Si Atom" : final_list
                        })
                try:
                        plt = sns.histplot(data=df, x="Frequency of Si Atom", kde=True,discrete=True,fill=False).set(title='Distribution of Si Atom in a Simulation Box')
                        fig = plt[0].get_figure()
                        fig.savefig("out_si.png")

                except Exception as e:
                        import traceback
                        print(traceback.format_exc())




#start = mimimum no of atoms per bucket, and end = maximum per bucket
def get_filtered_data(data,start,end):
        start = int(start)
        end = int(end)
        buckets = {}
        for i in range(NO_OF_BUCKET):
            buckets[i] = 0
        for each in data:
                datum = each['coordinate']
                x_cord, y_cord, z_cord = datum[0], datum[1], datum[2]
                x_offset = int((x_cord - LEFT_BOUNDARY)/BUCKET_LENGTH)
                y_offset = int((y_cord - LEFT_BOUNDARY)/BUCKET_LENGTH) 
                z_offset = int((z_cord - LEFT_BOUNDARY)/BUCKET_LENGTH)

                z_multiplier = round(NO_OF_BUCKET ** (2/3),2)
                y_multiplier = round(NO_OF_BUCKET ** (1/3),2)

                bucket_no = int((z_multiplier * z_offset + y_multiplier * y_offset + x_offset))
                try:
                    a = buckets[bucket_no]
                    a = a + 1
                    buckets[bucket_no] = a
                except Exception as e:
                    continue

        list_num_per_bucket = []
        """later"""
        return_buckets = []


        for x,y in buckets.items():
            list_num_per_bucket.append(y)
            if y >=start and y<=end:
                return_buckets.append(x)


        print("return_buckets--",return_buckets)

        out_file = open("filtered_bucket.json", "w")
          
        json.dump(return_buckets, out_file)
        out_file.close()
        filtered_data = []
        for obj in data:
                datum = obj['coordinate']
                datum_id = obj['id']
                x_cord, y_cord, z_cord = datum[0], datum[1], datum[2]
                x_offset = int((x_cord - LEFT_BOUNDARY)/BUCKET_LENGTH)
                y_offset = int((y_cord - LEFT_BOUNDARY)/BUCKET_LENGTH) 
                z_offset = int((z_cord - LEFT_BOUNDARY)/BUCKET_LENGTH)

                z_multiplier = round(NO_OF_BUCKET ** (2/3),2)
                y_multiplier = round(NO_OF_BUCKET ** (1/3),2)

                bucket_no = int((z_multiplier * z_offset + y_multiplier * y_offset + x_offset))
                if bucket_no in return_buckets:
                        # if x_cord > 0:
                        #       x_cord = x_cord - round(CELL_SIZE/2,2)
                        # else:
                        #       x_cord = x_cord + round(CELL_SIZE/2,2)
                        if y_cord > 0:
                                y_cord = y_cord - round(CELL_SIZE/2,2)
                        else:
                                y_cord = y_cord + round(CELL_SIZE/2,2)
                        # if z_cord > 0:
                        #       z_cord = z_cord - round(CELL_SIZE/2,2)
                        # else:
                        #       z_cord = z_cord + round(CELL_SIZE/2,2)
                        filtered_data.append(
                                {
                                    'id' : datum_id,
                                    'atom_coordinate' : [round(x_cord,2),round(y_cord,2),round(z_cord,2)] 
                                }
                            )

        if args.input == 'fe.json':
                out_file = open("filtered_fe.json", "w")
        elif args.input == 'mg.json':
                out_file = open("filtered_mg.json", "w")
        elif args.input == 'o.json':
                out_file = open("filtered_o.json", "w")
        else:
                out_file = open("filtered_si.json", "w")

        
        json.dump(filtered_data, out_file)
        out_file.close()


        # f = open("out_filtered.dump", "w")
        # f.write("ITEM: TIMESTEP\n")
        # f.write("0\n")
        # f.write("ITEM: NUMBER OF ATOMS\n")
        # f.write("{}\n".format(len(filtered_data)))
        # f.write("ITEM: BOX BOUNDS xy xz yz pp pp pp\n")
        # f.write("0.0000000000000000e+00 6.8000000000000000e+01 0.0000000000000000e+00\n")
        # f.write("0.0000000000000000e+00 6.8000000000000000e+01 0.0000000000000000e+00\n")
        # f.write("0.0000000000000000e+00 6.8000000000000000e+01 0.0000000000000000e+00\n")
        # f.write("ITEM: ATOMS id type x y z\n")



        # for each in filtered_data:
        #     f.write("{} {} {} {} {}".format(each['id'],"1",str(each['atom_coordinate'][0]),str(each['atom_coordinate'][1]),str(each['atom_coordinate'][2])))
        #     f.write("\n")
        # f.close()

def get_filtered_data_all(data,start,end,atom_class,sub_distribution=None):
        start = int(start)
        end = int(end)
        buckets = {}
        for i in range(NO_OF_BUCKET):
            buckets[i] = 0
        for each in data:
                if each['atom_class'] != atom_class:
                        continue
                datum = each['coordinate']
                x_cord, y_cord, z_cord = datum[0], datum[1], datum[2]
                x_offset = int((x_cord - LEFT_BOUNDARY)/BUCKET_LENGTH)
                y_offset = int((y_cord - LEFT_BOUNDARY)/BUCKET_LENGTH) 
                z_offset = int((z_cord - LEFT_BOUNDARY)/BUCKET_LENGTH)

                z_multiplier = round(NO_OF_BUCKET ** (2/3),2)
                y_multiplier = round(NO_OF_BUCKET ** (1/3),2)

                bucket_no = int((z_multiplier * z_offset + y_multiplier * y_offset + x_offset))
                try:
                    a = buckets[bucket_no]
                    a = a + 1
                    buckets[bucket_no] = a
                except Exception as e:
                    continue

        list_num_per_bucket = []
        """later"""
        # Return buckets contain all those bucket having threshold number of atoms ex: 35-50 Fe atom
        return_buckets = []

        for x,y in buckets.items():
            list_num_per_bucket.append(y)
            if y >=start and y<=end:
                return_buckets.append(x)


        print("return_buckets-",return_buckets)
        out_file = open("filtered_bucket.json", "w")
          
        json.dump(return_buckets, out_file)
        out_file.close()
        filtered_data = []
        filtered_data_before_translation = []
        for obj in data:
                datum = obj['coordinate']
                datum_id = obj['id']
                x_cord, y_cord, z_cord = datum[0], datum[1], datum[2]
                x_offset = int((x_cord - LEFT_BOUNDARY)/BUCKET_LENGTH)
                y_offset = int((y_cord - LEFT_BOUNDARY)/BUCKET_LENGTH) 
                z_offset = int((z_cord - LEFT_BOUNDARY)/BUCKET_LENGTH)

                z_multiplier = round(NO_OF_BUCKET ** (2/3),2)
                y_multiplier = round(NO_OF_BUCKET ** (1/3),2)

                bucket_no = int((z_multiplier * z_offset + y_multiplier * y_offset + x_offset))
                if bucket_no in return_buckets:
                        # if x_cord > 0:
                        #       x_cord = x_cord - round(CELL_SIZE/2,2)
                        # else:
                        #       x_cord = x_cord + round(CELL_SIZE/2,2)
                        if y_cord > 0:
                                y_cord_after_translation = y_cord - round(CELL_SIZE/2,2)
                        else:
                                y_cord_after_translation = y_cord + round(CELL_SIZE/2,2)
                        # if z_cord > 0:
                        #       z_cord = z_cord - round(CELL_SIZE/2,2)
                        # else:
                        #       z_cord = z_cord + round(CELL_SIZE/2,2)
                        filtered_data.append(
                                {
                                    'id' : datum_id,
                                    'atom_coordinate' : [round(x_cord,2),round(y_cord_after_translation,2),round(z_cord,2)],
                                    'atom_class' : obj['atom_class'] 
                                }
                            )
                        filtered_data_before_translation.append(
                                {
                                    'id' : datum_id,
                                    'atom_coordinate' : [round(x_cord,2),round(y_cord,2),round(z_cord,2)],
                                    'atom_class' : obj['atom_class'] 
                                }
                            )

        if atom_class == "1":
                out_file = open("cluster_within_fe.json", "w")
        elif atom_class == "2":
                out_file = open("cluster_within_mg.json", "w")
        elif atom_class == "3":
                out_file = open("cluster_within_o.json", "w")
        else:
                out_file = open("cluster_within_si.json", "w")

        
        json.dump(filtered_data, out_file)
        out_file.close()

        f_out = open("merged_cluster.dump", "w")
        f_out.write("ITEM: TIMESTEP\n")
        f_out.write("0\n")
        f_out.write("ITEM: NUMBER OF ATOMS\n")
        f_out.write("{}\n".format(str(len(filtered_data))))
        f_out.write("ITEM: BOX BOUNDS xy xz yz pp pp pp\n")
        f_out.write("0.0000000000000000e+00 6.8000000000000000e+01 0.0000000000000000e+00\n")
        f_out.write("0.0000000000000000e+00 6.8000000000000000e+01 0.0000000000000000e+00\n")
        f_out.write("0.0000000000000000e+00 6.8000000000000000e+01 0.0000000000000000e+00\n")
        f_out.write("ITEM: ATOMS id type x y z\n")


        for each in filtered_data:
                f_out.write("{} {} {} {} {}".format(each['id'],each['atom_class'],str(each['atom_coordinate'][0]),str(each['atom_coordinate'][1]),str(each['atom_coordinate'][2])))
                f_out.write("\n")

        f_out.close()
        get_atom_distribution(filtered_data_before_translation,filtering=True,atom_class=sub_distribution,return_buckets=return_buckets)




def merge_data(merge_val):
        total = 0 
        for i in merge_val:
                if i == "1":
                        f = open("filtered_fe.json","r")
                elif i == "2":
                        f = open("filtered_mg.json","r")
                elif i == "3":
                        f = open("filtered_o.json","r")
                else:
                        f = open("filtered_si.json","r")
                filtered_data = json.load(f)
                total = total + len(filtered_data)

        f_out = open("out_filtered.dump", "w")
        f_out.write("ITEM: TIMESTEP\n")
        f_out.write("0\n")
        f_out.write("ITEM: NUMBER OF ATOMS\n")
        f_out.write("{}\n".format(str(total)))
        f_out.write("ITEM: BOX BOUNDS xy xz yz pp pp pp\n")
        f_out.write("0.0000000000000000e+00 6.8000000000000000e+01 0.0000000000000000e+00\n")
        f_out.write("0.0000000000000000e+00 6.8000000000000000e+01 0.0000000000000000e+00\n")
        f_out.write("0.0000000000000000e+00 6.8000000000000000e+01 0.0000000000000000e+00\n")
        f_out.write("ITEM: ATOMS id type x y z\n")
        for i in merge_val:
                if i == "1":
                        f = open("filtered_fe.json","r")
                elif i == "2":
                        f = open("filtered_mg.json","r")
                elif i == "3":
                        f = open("filtered_o.json","r")
                else:
                        f = open("filtered_si.json","r")
                filtered_data = json.load(f)

                for each in filtered_data:
                        f_out.write("{} {} {} {} {}".format(each['id'],str(i),str(each['atom_coordinate'][0]),str(each['atom_coordinate'][1]),str(each['atom_coordinate'][2])))
                        f_out.write("\n")
        f_out.close()

def calculate_distance_from_distance_matrix(distance_matrix,si_id,fe_id):
        if si_id < fe_id:
                distance = distance_matrix[si_id][fe_id]
        else:
                distance = distance_matrix[fe_id][si_id]

        return distance

def get_proximity(data,skip=True):
        import json
        if not skip:
                import math
                distance_matrix = {}
                atom_class_id = {
                        '1' : [],
                        '2' : [],
                        '3' : [],
                        '4' : []

                }
                for i,first in enumerate(data):
                        distance_matrix[first["id"]] = {}
                        temp = distance_matrix[first["id"]]

                        """Start Atom id for each class"""
                        if first['atom_class'] == "1":
                                atom_class_id['1'].append(first["id"])
                        elif first['atom_class'] == "2":
                                atom_class_id['2'].append(first["id"])
                        elif first['atom_class'] == "3":
                                atom_class_id['3'].append(first["id"])
                        else:
                                atom_class_id['4'].append(first["id"])
                        """End Atom id for each class"""

                        for second in data[i+1:]:
                                distance_map = {}
                                a1 = first['atom_coordinate']
                                b1 = second['atom_coordinate']
                                distance = round(math.sqrt((a1[0]-b1[0])**2 + (a1[1]-b1[1])**2 + (a1[2]-b1[2])**2),2)
                                temp[second["id"]] = distance

                out_file = open("distance_matrix.json", "w")
                import json
                json.dump(distance_matrix,out_file)
                out_file.close()

                out_file_2 = open("atom_class_id.json","w")
                json.dump(atom_class_id,out_file_2)
                out_file_2.close()
        else:
                with open('distance_matrix.json','r') as myfile:
                        distance_matrix = json.load(myfile)

                myfile.close()
                with open('atom_class_id.json') as file:
                        atom_class_id = json.load(file)

                file.close()

                CUT_OFF = 2.5

                si_proximity_count = {}
                # o_proximity_count = {}
                for si_id in atom_class_id["3"]:
                        si_proximity_count[si_id] = {
                                "Fe" : 0,
                                "Si" : 0,
                                "O" : 0,
                                "Mg" : 0
                        }
                        for fe_id in atom_class_id["1"]:
                                distance = calculate_distance_from_distance_matrix(distance_matrix,si_id,fe_id)
                                if distance <= CUT_OFF:
                                        si_proximity_count[si_id]["Fe"] += 1

                        for mg_id in atom_class_id["2"]:
                                distance = calculate_distance_from_distance_matrix(distance_matrix,si_id,mg_id)
                                if distance <= CUT_OFF:
                                        si_proximity_count[si_id]["Mg"] += 1

                        for s2_id in atom_class_id["3"]:
                                if si_id == s2_id:
                                        continue
                                distance = calculate_distance_from_distance_matrix(distance_matrix,si_id,s2_id)
                                if distance <= CUT_OFF:
                                        si_proximity_count[si_id]["Si"] += 1

                        for o_id in atom_class_id["4"]:
                                distance = calculate_distance_from_distance_matrix(distance_matrix,si_id,o_id)
                                if distance <= CUT_OFF:
                                        si_proximity_count[si_id]["O"] += 1


                # for o_id in atom_class_id["4"]:
                #       o_proximity_count[o_id] = {
                #               "Fe" : 0,
                #               "Si" : 0,
                #               "O" : 0,
                #               "Mg" : 0
                #       }
                #       for fe_id in atom_class_id["1"]:
                #               distance = calculate_distance_from_distance_matrix(distance_matrix,o_id,fe_id)
                #               if distance <= CUT_OFF:
                #                       o_proximity_count[o_id]["Fe"] += 1

                #       for mg_id in atom_class_id["2"]:
                #               distance = calculate_distance_from_distance_matrix(distance_matrix,o_id,mg_id)
                #               if distance <= CUT_OFF:
                #                       o_proximity_count[o_id]["Mg"] += 1

                #       for s2_id in atom_class_id["3"]:
                #               distance = calculate_distance_from_distance_matrix(distance_matrix,o_id,s2_id)
                #               if distance <= CUT_OFF:
                #                       o_proximity_count[o_id]["Si"] += 1

                #       for o2_id in atom_class_id["4"]:
                #               if o2_id == o_id:
                #                       continue
                #               distance = calculate_distance_from_distance_matrix(distance_matrix,o_id,o2_id)
                #               if distance <= CUT_OFF:
                #                       o_proximity_count[o_id]["O"] += 1

                
                # out_file = open("silicon_proximity.json", "w")
                # import json
                # json.dump(si_proximity_count,out_file)
                # out_file.close()
                # print("Outputing Silicon Proximity count as silicon_proximity.json")

                # fe_count = []
                # si_count = []
                # mg_count = []
                # o_count = []
                # for si_id, data in si_proximity_count.items():
                #       fe_count.append(data["Fe"])
                #       mg_count.append(data["Mg"])
                #       si_count.append(data["Si"])
                #       o_count.append(data["O"])

                
                # out_file_1 = open("fe_count_in_proximity.json", "w")
                # json.dump(fe_count,out_file_1)
                # out_file_1.close()

                # out_file_2 = open("mg_count_in_proximity.json", "w")
                # json.dump(mg_count,out_file_2)
                # out_file_2.close()

                # out_file_3 = open("si_count_in_proximity.json", "w")
                # json.dump(si_count,out_file_3)
                # out_file_3.close()

                # out_file_4 = open("o_count_in_proximity.json", "w")
                # json.dump(o_count,out_file_4)
                # out_file_4.close()


                """Start count for oxygen for histogram"""
                out_file_2 = open("silicon_proximity.json","w")
                json.dump(si_proximity_count,out_file_2)
                out_file_2.close()
                print("Outputing Silicon Proximity count as silicon_proximity.json")


                fe_count = []
                si_count = []
                mg_count = []
                o_count = []
                for o_id, data in si_proximity_count.items():
                        fe_count.append(data["Fe"])
                        mg_count.append(data["Mg"])
                        si_count.append(data["Si"])
                        o_count.append(data["O"])

                out_file_1 = open("fe_count_in_proximity.json", "w")
                json.dump(fe_count,out_file_1)
                out_file_1.close()

                out_file_2 = open("mg_count_in_proximity.json", "w")
                json.dump(mg_count,out_file_2)
                out_file_2.close()

                out_file_3 = open("si_count_in_proximity.json", "w")
                json.dump(si_count,out_file_3)
                out_file_3.close()

                out_file_4 = open("o_count_in_proximity.json", "w")
                json.dump(o_count,out_file_4)
                out_file_4.close()

                """End count for oxygen for histogram"""

def tsne(data):
        import numpy as np
        import pandas as pd
        import matplotlib.pyplot as plt
        import seaborn as sns
        from sklearn.manifold import TSNE
        from sklearn.preprocessing import StandardScaler

        df_dict = {
        'id' : [],
        'x' : [],
        'y': [],
        'z': [],
        'class' : []
        }

        for datum in data:
                df_dict['x'].append(round(datum['coordinate'][0],2))
                df_dict['y'].append(round(datum['coordinate'][1],2))
                df_dict['z'].append(round(datum['coordinate'][2],2))
                df_dict['class'].append(datum['atom_class'])
                df_dict['id'].append(datum['id'])

        import pandas as pd

        df = pd.DataFrame(df_dict)

        df.to_csv('data3.csv',index=False)

        data = df.iloc[:,1:-1].to_numpy()
        label = df.iloc[:,-1].to_numpy()

        model = TSNE(n_components = 2, random_state = 0)
        tsne_data = model.fit_transform(data)
        tsne_data = np.vstack((tsne_data.T, label)).T
        tsne_df = pd.DataFrame(data = tsne_data,
     columns =("Dim1", "Dim2", "label"))


        sns.FacetGrid(tsne_df, hue ="label").map(
        plt.scatter, 'Dim1', 'Dim2').add_legend()
 
        plt.show()

def atom_movement_data(all_data,atom_class):
        import math
        sign_map = {}
        dict_data = {}
        dict_data['atom_id'] = []
        for i,each_sim_data in enumerate(all_data):
                dict_data[i+1] = []
                for each in each_sim_data:
                        if each['atom_class'] == atom_class:
                                if i==0: #can be done only once for 1st simulation data
                                        dict_data['atom_id'].append(each['id'])
                                a1 = each['coordinate']
                                b1 = [0,0,0]
                                distance = round(math.sqrt((a1[0]-b1[0])**2 + (a1[1]-b1[1])**2 + (a1[2]-b1[2])**2),2)

                                dict_data[i+1].append(distance)

        import pandas as pd
        plot_x = []
        plot_y = []
        for x,y in dict_data.items():
                if x=='atom_id':
                        continue
                plot_x.append(int(x))
                plot_y.append(sum(y)/len(y))


        from matplotlib import  pyplot as plt
        plt.plot(plot_x,plot_y)
        plt.show()








# # TSNE
# # Picking the top 1000 points as TSNE
# # takes a lot of time for 15K points
# data_1000 = standardized_data[0:1000, :]
# labels_1000 = labels[0:1000]
 
# model = TSNE(n_components = 2, random_state = 0)
# # configuring the parameters
# # the number of components = 2
# # default perplexity = 30
# # default learning rate = 200
# # default Maximum number of iterations
# # for the optimization = 1000
 
# tsne_data = model.fit_transform(data_1000)
 
# # creating a new data frame which
# # help us in plotting the result data
# tsne_data = np.vstack((tsne_data.T, labels_1000)).T
# tsne_df = pd.DataFrame(data = tsne_data,
#      columns =("Dim_1", "Dim_2", "label"))
 
# # Plotting the result of tsne
# sn.FacetGrid(tsne_df, hue ="label", size = 6).map(
#        plt.scatter, 'Dim_1', 'Dim_2').add_legend()
 
# plt.show()


def scatterplot():
        with open('buckets_fe_density.json', 'r') as fe:
            fe_data = json.load(fe)

        with open('buckets_mg_density.json', 'r') as mg:
            mg_data = json.load(mg)

        with open('buckets_si_density.json', 'r') as si:
            si_data = json.load(si)

        with open('buckets_o_density.json', 'r') as o:
            o_data = json.load(o)


        fe_data_list = [y for x,y in fe_data.items()]
        mg_data_list = [y for x,y in mg_data.items()]
        si_data_list = [y for x,y in si_data.items()]
        o_data_list = [y for x,y in o_data.items()]

        zipped_val = zip(fe_data_list,mg_data_list,si_data_list,o_data_list)
        import pandas as pd
        df = pd.DataFrame(zipped_val,columns=['Fe','Mg','Si','O'])
        corr = df.corr(min_periods=3)
        print(corr)
        ax1 = df.plot(kind='scatter', x='Fe', y='Mg', color='r') 


def parsed_data_xdat_car(x,y,z,i):
        print("x=",x,y,z)
        size = int(args.cell_size)
        output = args.output

        # dict_val = dict(zip_info)
        # dict_val = {
        # 'Fe' : 85,
        # 'Mg' : 104
        # }


        ATOM_CLASS_MAPPER_2 = {
                'fe' : '1',
                'mg' : '2',
                'si' : '3',
                'o' : '4',
                'h' : '5',
                'n' : '5',
                'c' : '5'
        }

        # if ATOM_CLASS_MAPPER_2[output]:
        #       pass


        return {
                'atom_id' : str(i),
                'atom_class' : ATOM_CLASS_MAPPER_2[output],
                'atom_coordinate' : [round(float(x)*size,2),round(float(y)*size,2),round(float(z)*size,2)]
        }

def xdatcar_to_json(all_data,path):
        file = open(path,"r")
        read_it = file.read()

        lines = read_it.splitlines()
        flag = 0
        final_data = []
        one_snap = []
        count = 0
        for i,line in enumerate(lines):
                if i < OFFSET_DISTANCE:
                        if i == 5:
                                atoms = line.strip().split(" ")
                                atoms = [val for val in atoms if val]
                        elif i == 6:
                                vals = line.strip().split(" ")
                                vals = [val for val in vals if val]
                        continue
                else:
                        if i < OFFSET_DISTANCE + 1:
                                #do it only once
                                zip_info = dict(zip(atoms,vals))
                                zip_required = {}
                                previous = 1
                                for x,y in zip_info.items():
                                    zip_required[x] = (previous,previous + int(float(y)) - 1)
                                    previous = previous + int(float(y))

                        if "Direct configuration" in line:
                                count = 0
                                if flag:
                                        final_data.append(one_snap)
                                        one_snap = []
                                flag = 1
                        else:
                                count = count + 1
                                if flag:
                                        print(zip_required)
                                        op = args.output
                                        if op == "fe":
                                                if not (count >= zip_required['Fe'][0] and count <= zip_required['Fe'][1]):
                                                        continue
                                        elif op == "mg":
                                                if not (count >= zip_required['Mg'][0] and count <= zip_required['Mg'][1]):
                                                        continue
                                        elif op == "si":
                                                if not (count >= zip_required['Si'][0] and count <= zip_required['Si'][1]):
                                                        continue
                                        elif op == "o":
                                                if not (count >= zip_required['O'][0] and count <= zip_required['O'][1]):
                                                        continue
                                        elif op == "n":
                                                if not (count >= zip_required['N'][0] and count <= zip_required['N'][1]):
                                                        continue
                                        elif op == "h":
                                                if not (count >= zip_required['H'][0] and count <= zip_required['H'][1]):
                                                        continue
                                        elif op == "c":
                                                if not (count >= zip_required['C'][0] and count <= zip_required['C'][1]):
                                                        continue
                                        line = line.strip()
                                        print("line=-",line)
                                        match = re.match("([0-9]*\.*\d*)[ ]*([0-9]*\.*\d*)[ ]*([0-9]*\.*\d*)",line)
                                        if match:
                                                print(match.group(1),match.group(2),match.group(2))
                                                val = parsed_data_xdat_car(match.group(1),match.group(2),match.group(3),count)
                                                one_snap.append(val)
        if one_snap:#for the last snapshot
                final_data.append(one_snap)

        out_file = open("{}.json".format(args.output), "w")

        json.dump(final_data, out_file, indent = 6)


def get_atom_distribution_average(all_data):
        import pandas as pd
        buckets = {}
        no_of_samples = 0
        my_distro = {}
        for i in range(MAX_ATOMS_PER_BUCKET+1):
            my_distro[i] = 0

        for i,each_data in enumerate(all_data):
                print(i)
                if not 1:
                        continue
                else:
                        no_of_samples = no_of_samples + 1
                for i in range(NO_OF_BUCKET):
                    buckets[i] = 0
                for datum in each_data:
                        atom_coordinate = datum['coordinate']
                        x_cord, y_cord, z_cord = atom_coordinate[0], atom_coordinate[1], atom_coordinate[2]
                        x_offset = int((x_cord - LEFT_BOUNDARY)/BUCKET_LENGTH)
                        y_offset = int((y_cord - LEFT_BOUNDARY)/BUCKET_LENGTH) 
                        z_offset = int((z_cord - LEFT_BOUNDARY)/BUCKET_LENGTH)

                        z_multiplier = round(NO_OF_BUCKET ** (2/3),2)
                        y_multiplier = round(NO_OF_BUCKET ** (1/3),2)

                        bucket_no = int((z_multiplier * z_offset + y_multiplier * y_offset + x_offset))
                        try:
                            a = buckets[bucket_no]
                            a = a + 1
                            buckets[bucket_no] = a
                        except Exception as e:
                            continue
                print("buckets=",buckets)
                print("my distro-",my_distro)
                list_num_per_bucket = []
                for x,y in buckets.items():
                    list_num_per_bucket.append(y)

                unique_elements = list(set(list_num_per_bucket))
                print("unique_elements=",unique_elements)
                print("list_num_per_bucket=",list_num_per_bucket)
                for element in unique_elements:
                        print("--------------------------------")
                        count = countX(list_num_per_bucket,element)
                        if element == 7:
                                print("Element is 7 count is {}".format(count))
                        my_distro[element] = my_distro[element] + count


        final_frequency = {}
        for key,value in my_distro.items():
            final_frequency[key] = round(value / no_of_samples,2)

        print("final Frequency-",final_frequency)
        temp = []
        for x,y in final_frequency.items():
                temp.append(y)
        print("final Frequency list-",temp)


        data = []
        for x,y in final_frequency.items():
            data = data + int(round(y,2)) * [x]

        final_list = []
        for i in data:
            final_list.append(int(round(i,2)))
        graph_output = args.output
        df = pd.DataFrame({"Frequency of {} Atom".format(ATOM_CLASS_MAPPER[graph_output]) : final_list})
        print(df.head())
        try:
                import seaborn as sns
                plt2 = sns.histplot(data=df, x="Frequency of {} Atom".format(ATOM_CLASS_MAPPER[graph_output]), kde=True,discrete=True,fill=False)
                plt2.set(title='Distribution of {} Atom in a Simulation Box'.format(ATOM_CLASS_MAPPER[graph_output]))
                plt2.bar_label(plt2.containers[0])
                fig2 = plt2.get_figure()
                fig2.savefig("outfinal_{}.png".format(ATOM_CLASS_MAPPER[graph_output]))
                print("zz")
        except Exception as e:
                print("ss")
                import traceback
                print(traceback.format_exc())


        # Report lab 

        from reportlab.platypus import Paragraph
        from reportlab.platypus import SimpleDocTemplate, Image
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from io import BytesIO

            
        # importing os module  
        import os 
          
        # Get the list of all files and directories 
        # in the root directory 
        path = "."
        dir_list = os.listdir(path)

        #Get the last analysis_{} val
        analysis_list_max_val = max([int(dir.split("_")[1].split(".")[0]) if "AnalysisNumber_" in dir else 0 for dir in dir_list])

        sample_style_sheet = getSampleStyleSheet()
        yourStyle = ParagraphStyle('yourtitle',
                           fontName="Helvetica",
                           fontSize=16,
                           parent=sample_style_sheet['BodyText'],
                           alignment=1,
                           spaceAfter=14,
                           leading=20
                           )

        flowables = []
        xlsx_file = "temp.xlsx"
        """
        if analysis_list_max_val == 0:
                my_doc = SimpleDocTemplate('AnalysisNumber_1.pdf')
        else:
                my_doc = SimpleDocTemplate('AnalysisNumber_{}.pdf'.format(str(analysis_list_max_val+1)))
                # xlsx_file = "temp_{}.xlsx".format(str(analysis_list_max_val+1))
        """
        my_doc = SimpleDocTemplate('AnalysisNumber_{}.pdf'.format(ATOM_CLASS_MAPPER_NUMBER[graph_output]))
        # pdf_buffer = BytesIO()
        # my_doc = SimpleDocTemplate(pdf_buffer)
        paragraph_1 = Paragraph("DPMD Analysis", sample_style_sheet['Heading1'])
        paragraph_2 = Paragraph("{} atom  distribution".format(ATOM_CLASS_MAPPER[graph_output]), sample_style_sheet['Heading1'])
        paragraph_3 = Paragraph(
            str(temp),
            yourStyle
        )

        image = Image("outfinal_{}.png".format(ATOM_CLASS_MAPPER[graph_output]))
        flowables.append(paragraph_1)
        flowables.append(paragraph_2)
        flowables.append(paragraph_3)
        flowables.append(image)
        # flowables.append(paragraph_3)

        my_doc.build(flowables)
        
        from openpyxl import Workbook
        import os
        if not os.path.exists(xlsx_file):
            wb = Workbook()
            ws = wb.active
            ws.cell(row=1,column=1,value="Bin Count")
            json_files = [f.split(".")[0] for f in os.listdir() if f.endswith(".json")]
            for cell in json_files:
                if cell in ["fe","mg","si","o","n","h"]:
                    ws.cell(row=1,column=ATOM_CLASS_MAPPER_NUMBER[cell]+1,value="{} cell".format(ATOM_CLASS_MAPPER[cell]))
                    ws.cell(row=1,column=ATOM_CLASS_MAPPER_NUMBER[cell]+8,value="{} cell (%)".format(ATOM_CLASS_MAPPER[cell]))
        
            for bin_number, value in enumerate(temp):
                ws.cell(row=bin_number+2, column=1, value=bin_number)
                ws.cell(row=bin_number+2, column=ATOM_CLASS_MAPPER_NUMBER[graph_output] + 1, value=value)
                ws.cell(row=bin_number+2, column=ATOM_CLASS_MAPPER_NUMBER[graph_output] + 8, value=round((value/NO_OF_BUCKET)*100,2))
        else:
            from openpyxl import load_workbook
            wb = load_workbook(xlsx_file)
            ws = wb.active
            for bin_number, value in enumerate(temp):
                ws.cell(row=bin_number+2, column=ATOM_CLASS_MAPPER_NUMBER[graph_output] + 1, value=value)
                ws.cell(row=bin_number+2, column=ATOM_CLASS_MAPPER_NUMBER[graph_output] + 8, value=round((value/NO_OF_BUCKET)*100,2))
        
        wb.save(xlsx_file)
        """
        if analysis_list_max_val == 0:
                print("1")
                from openpyxl import Workbook
                
                wb = Workbook()
                ws = wb.active
                
                ws.cell(row=1,column=
                #ws.append(["Bin Count","Frequency {}".format(ATOM_CLASS_MAPPER[graph_output])])
                ws.cell(row=1, column=ATOM_CLASS_MAPPER_NUMBER[graph_output] + 8, 
        value=f"{ATOM_CLASS_MAPPER[graph_output]} cell %")
                for bin_number, value in enumerate(temp):
                        ws.append([bin_number, value])
                        ws.cell(row=bin_number+2, column=ATOM_CLASS_MAPPER_NUMBER[graph_output] + 8, value=round((value/NO_OF_BUCKET)*100,2))
                wb.save(xlsx_file)

        else:
                print("2")
                print(ATOM_CLASS_MAPPER_NUMBER[graph_output] + 1)
                print(graph_output)
                from openpyxl import load_workbook
                wb = load_workbook(xlsx_file)
                ws = wb.active
                ws.cell(row=1, column=ATOM_CLASS_MAPPER_NUMBER[graph_output] + 1, value="Frequency {}".format(ATOM_CLASS_MAPPER[graph_output]))
                ws.cell(row=1, column=ATOM_CLASS_MAPPER_NUMBER[graph_output] + 8, value="{} cell %".format(ATOM_CLASS_MAPPER[graph_output]))
                for row, value in enumerate(temp, start=2):  # Start from row 2 (after the header)
                        ws.cell(row=row, column=ATOM_CLASS_MAPPER_NUMBER[graph_output] + 1, value=value)
                        ws.cell(row=row, column=ATOM_CLASS_MAPPER_NUMBER[graph_output] + 8, value=round((value/NO_OF_BUCKET)*100,2))
                wb.save(xlsx_file)
            """
def get_sub_atom_count_in_fe(all_data_fe,all_data_n):
        print("length-",len(all_data))
        min_count = 0
        max_count = MAX_ATOMS_PER_BUCKET
        final_data = {}
        for i in range(min_count,max_count+1):
                final_data[i] = []

        for k, fe_data in enumerate(all_data_fe):
                buckets = {}
                for i in range(NO_OF_BUCKET):
                    buckets[i] = 0
                for each in fe_data:
                        datum = each['coordinate']
                        x_cord, y_cord, z_cord = datum[0], datum[1], datum[2]
                        x_offset = int((x_cord - LEFT_BOUNDARY)/BUCKET_LENGTH)
                        y_offset = int((y_cord - LEFT_BOUNDARY)/BUCKET_LENGTH) 
                        z_offset = int((z_cord - LEFT_BOUNDARY)/BUCKET_LENGTH)

                        z_multiplier = round(NO_OF_BUCKET ** (2/3),2)
                        y_multiplier = round(NO_OF_BUCKET ** (1/3),2)

                        bucket_no = int((z_multiplier * z_offset + y_multiplier * y_offset + x_offset))
                        try:
                            a = buckets[bucket_no]
                            a = a + 1
                            buckets[bucket_no] = a
                        except Exception as e:
                            continue

                count_buckets = {}
                for i in range(min_count,max_count+1):
                        count_buckets[i] = []

                for x,y in buckets.items():
                        # count_buckets key value contains lists of all the buckets that contain 0 atoms or 1 atoms or 2 atoms or 8atoms.
                        count_buckets[y].append(x)

                # print("FOr k ={}, count buckets={}".format(k,count_buckets))
                # Eg:
                # {

                #       0 : [1,4,12,64],
                #       1: [2,3,5,7,8,9,63],2
                #       2: [....]
                # }


                # for 1st instance of Fe atom, we check the first instance of N atom
                n_data = all_data_n[k]
                # Contains number of nitrogen count in n count Fe bin
                # print("count of sub atom={}".format(len(n_data)))
                count_nitrogen_in_fe_count = {}
                for i in range(min_count,max_count+1):
                        count_nitrogen_in_fe_count[i] = 0
                for each in n_data:
                        datum = each['coordinate']
                        x_cord, y_cord, z_cord = datum[0], datum[1], datum[2]
                        x_offset = int((x_cord - LEFT_BOUNDARY)/BUCKET_LENGTH)
                        y_offset = int((y_cord - LEFT_BOUNDARY)/BUCKET_LENGTH) 
                        z_offset = int((z_cord - LEFT_BOUNDARY)/BUCKET_LENGTH)

                        z_multiplier = round(NO_OF_BUCKET ** (2/3),2)
                        y_multiplier = round(NO_OF_BUCKET ** (1/3),2)

                        bucket_no = int((z_multiplier * z_offset + y_multiplier * y_offset + x_offset))
                        # print("bucket_no=",bucket_no)
                        for x,y in count_buckets.items():
                                if bucket_no in y:
                                        count_nitrogen_in_fe_count[x]+=1


                for a,b in count_nitrogen_in_fe_count.items():
                        final_data[a].append(b)

        import pandas as pd
        df = pd.DataFrame(final_data)

        df.loc['Average'] = df.mean()
        df.loc['Max'] = df.max()
        df.loc['Min'] = df.min()

        df = df.reindex(np.roll(df.index, shift=3))

        df.to_excel("{}_Count.xlsx".format(args.output)) 


def get_connected_componen(data,start,end):

        start = int(start)
        end = int(end)
        buckets = {}
        for i in range(NO_OF_BUCKET):
            buckets[i] = 0
        for each in data:
                datum = each['coordinate']
                x_cord, y_cord, z_cord = datum[0], datum[1], datum[2]
                x_offset = int((x_cord - LEFT_BOUNDARY)/BUCKET_LENGTH)
                y_offset = int((y_cord - LEFT_BOUNDARY)/BUCKET_LENGTH) 
                z_offset = int((z_cord - LEFT_BOUNDARY)/BUCKET_LENGTH)

                z_multiplier = round(NO_OF_BUCKET ** (2/3),2)
                y_multiplier = round(NO_OF_BUCKET ** (1/3),2)

                bucket_no = int((z_multiplier * z_offset + y_multiplier * y_offset + x_offset))
                try:
                    a = buckets[bucket_no]
                    a = a + 1
                    buckets[bucket_no] = a
                except Exception as e:
                    continue

        list_num_per_bucket = []
        """later"""

        #These are the buckets with start-end (5-8) no of atoms/ high density buckets/ Fe rich
        return_buckets = []


        for x,y in buckets.items():
            list_num_per_bucket.append(y)
            if y >=start and y<=end:
                return_buckets.append(x)

        return_buckets_coordiate = {}
        k = NO_OF_BUCKET ** (1/3)
        for num in return_buckets:
            x = int((num - 1)/k**2) + 1
            y = int(((num - (x-1)*k**2) - 1)/k) + 1
            z = (num-1) - (k**2*(x-1)) - (k*(y-1)) + 1
            return_buckets_coordiate[num] = [round(x,2),round(y,2),round(z,2)]
        k = round(k,1)
        graph = {}
        import numpy as np
        from collections import Counter
        import networkx as nx
        for x,y in return_buckets_coordiate.items():
                graph[x] = []
                possible_values_of_delta = [1]
                for a,b in return_buckets_coordiate.items():
                        if x == a:
                                continue
                        diff = abs(np.array(y) - np.array(b))
                        count_map = Counter(list(diff))
                        
                        #For border bins
                        if y[0] in [1,k] or y[1] in [1,k] or y[2] in [1,k]:
                                if y[0] in [1,k]: # border along x -direction | only allow delta = (1,0,0) and (k-1,0,0) to be connected

                                        if ((count_map.get(1,0) == 1 or count_map.get(k-1) ==1) and count_map.get(0,0) ==2): #only allow delta = (1,0,0) and (k-1,0,0) to be connected
                                                graph[x].append(a)

                                if y[1] in [1,k]: # border along y -direction | only allow delta = (1,0,0) and (k-1,0,0) to be connected
                                        diff = abs(np.array(y) - np.array(b))
                                        count_map = Counter(list(diff))
                                        if ((count_map.get(1,0) == 1 or count_map.get(k-1) ==1) and count_map.get(0,0) ==2): #only allow delta = (0,1,0) and (0,k-1,0) to be connected
                                                graph[x].append(a)

                                if y[2] in [1,k]: # border along z -direction | only allow delta = (0,0,1) and (0,0,k-1) to be connected
                                        diff = abs(np.array(y) - np.array(b))
                                        count_map = Counter(list(diff))
                                        if ((count_map.get(1,0) == 1 or count_map.get(k-1) ==1) and count_map.get(0,0) ==2): #only allow delta = (1,0,0) and (k-1,0,0) to be connected
                                                graph[x].append(a)

                        #For middle bins
                        else:
                                # diff of two cell cordinates must be in [[1,0,0],......[0,0,1]]
                                diff = abs(np.array(y) - np.array(b))
                                count_map = Counter(list(diff))

                                if count_map.get(1,0) == 1 and count_map.get(0,0) == 2 or count_map.get(1,0) == 2 and count_map.get(0,0) == 1:
                                        graph[x].append(a)

        
        unique_graph = {}
        for x,y in graph.items():
                unique_graph[x] = list(set(graph[x]))
        print("unique_elements-",unique_graph)
        G = nx.Graph(unique_graph)
        final_components = []
        for i in nx.connected_components(G):
                final_components.append(i)
        print("Connected components are : {}".format(final_components))

def alpha_shape_3D(pos, alpha):
        """
        Compute the alpha shape (concave hull) of a set of 3D points.
        Parameters:
            pos - np.array of shape (n,3) points.
            alpha - alpha value.
        return
            outer surface vertex indices, edge indices, and triangle indices
        """
        from scipy.spatial import Delaunay
        from collections import defaultdict
        tetra = Delaunay(pos)
        # Find radius of the circumsphere.
        # By definition, radius of the sphere fitting inside the tetrahedral needs 
        # to be smaller than alpha value
        # http://mathworld.wolfram.com/Circumsphere.html
        tetrapos = np.take(pos,tetra.vertices,axis=0)
        normsq = np.sum(tetrapos**2,axis=2)[:,:,None]
        ones = np.ones((tetrapos.shape[0],tetrapos.shape[1],1))
        a = np.linalg.det(np.concatenate((tetrapos,ones),axis=2))
        Dx = np.linalg.det(np.concatenate((normsq,tetrapos[:,:,[1,2]],ones),axis=2))
        Dy = -np.linalg.det(np.concatenate((normsq,tetrapos[:,:,[0,2]],ones),axis=2))
        Dz = np.linalg.det(np.concatenate((normsq,tetrapos[:,:,[0,1]],ones),axis=2))
        c = np.linalg.det(np.concatenate((normsq,tetrapos),axis=2))
        r = np.sqrt(Dx**2+Dy**2+Dz**2-4*a*c)/(2*np.abs(a))

        # Find tetrahedrals
        tetras = tetra.vertices[r<alpha,:]
        # triangles
        TriComb = np.array([(0, 1, 2), (0, 1, 3), (0, 2, 3), (1, 2, 3)])
        Triangles = tetras[:,TriComb].reshape(-1,3)
        Triangles = np.sort(Triangles,axis=1)
        # Remove triangles that occurs twice, because they are within shapes
        TrianglesDict = defaultdict(int)
        for tri in Triangles:TrianglesDict[tuple(tri)] += 1
        Triangles=np.array([tri for tri in TrianglesDict if TrianglesDict[tri] ==1])
        #edges
        EdgeComb=np.array([(0, 1), (0, 2), (1, 2)])
        Edges=Triangles[:,EdgeComb].reshape(-1,2)
        Edges=np.sort(Edges,axis=1)
        Edges=np.unique(Edges,axis=0)

        Vertices = np.unique(Edges)
        return Vertices,Edges,Triangles

def area_of_triangle(p1,p2,p3):
        import math
        p1 = np.array(p1)
        p2 = np.array(p2)
        p3 = np.array(p3)

        v1 = p2-p1
        v2 = p3-p1

        v = np.cross(v1,v2)
        area = (1/2) * math.sqrt(math.pow(v[0],2) + math.pow(v[1],2) + math.pow(v[2],2))
        return area

def is_within_triangle(p1,p2,p3,p):

    a1 = area_of_triangle(p,p1,p2)
    a2 = area_of_triangle(p,p1,p3)
    a3 = area_of_triangle(p,p2,p3)

    a = area_of_triangle(p1,p2,p3)

    if abs((a1 + a2 + a3) - a) < 0.1:
        return True
    else:
        return False

def atom_count_in_alpha_shapes(data,input_data,n_data):
        if type(input_data[0]) == list:
                n_data = int(round(n_data,2))
                input_data = input_data[n_data]
        required_data = []
        for datum in data:
                required_data.append(datum['atom_coordinate'])

        pts = np.array(required_data)
        verts, edges, faces = alpha_shape_3D(pts, alpha=2)
        initial_verts = verts.copy()
        initial_verts = list(initial_verts)
        verts = np.array([list(pts[index]) for index in verts])
        list2 = []
        for inner in faces:
            list_each = list(inner)
            altered_list = [initial_verts.index(x) for x in list_each ]
            list2.append(np.array(altered_list))
        faces = np.array(list2)
        #count start for each input data
        outside_count = 0
        inside_count = 0
        # print("vert=",verts)
        # print("input_data0=",input_data)
        for each_atom in input_data:
                ips = []
                p = each_atom['atom_coordinate'][0]
                q = each_atom['atom_coordinate'][1]
                r = each_atom['atom_coordinate'][2]
                for i, face in enumerate(faces):
                    x1, y1, z1 = verts[face[0]][0],verts[face[0]][1],verts[face[0]][2]
                    x2, y2, z2 = verts[face[1]][0],verts[face[1]][1],verts[face[1]][2]
                    x3, y3, z3 = verts[face[2]][0],verts[face[2]][1],verts[face[2]][2]

                    x_face_max = max(x1,x2,x3)
                    y_face_max = max(y1,y2,y3)
                    z_face_max = max(z1,z2,z3)

                    p1 = np.array([x1,y1,z1])
                    p2 = np.array([x2,y2,z2])
                    p3 = np.array([x3,y3,z3])

                    v1 = p3 - p1
                    v2 = p2 - p1
                    cp = np.cross(v1, v2)

                    a, b, c = cp
                    d = np.dot(cp, p3)
                    if a==0:
                        a = 0.00001
                    t = round((-1* a*p - b*q - c*r + d)/ a,2)

                    x = p + t
                    y = q
                    z = r
                    if 1:
                        # if x <= p and is_within_triangle([x1,y1,z1],[x2,y2,z2],[x3,y3,z3],[x,y,z]):
                        #     ips.append([x,y,z])
                        # if not x<=p and is_within_triangle([x1,y1,z1],[x2,y2,z2],[x3,y3,z3],[x,y,z]):
                        #       print("ayoo=",[x,y,z])
                        #       print("p=",p)
                        if x>p and is_within_triangle([x1,y1,z1],[x2,y2,z2],[x3,y3,z3],[x,y,z]):
                            ips.append([x,y,z])



                # if len(ips) % 2==0:
                #       outside_count += 1
                #       print("outside_count=",outside_count)
                #       print("ips=",ips)
                # else:
                #       inside_count+=1
                #       print("inside_count=",inside_count)
                if [p,q,r] in verts:
                        inside_count+=1
                        continue
                if len(ips) % 2 == 0 and len(ips)!=0:
                        # this is for points in common edge
                        if len(ips) == 2 and ips[0][0] == ips[1][0]:
                                inside_count += 1
                        else:
                                outside_count += 1
                elif len(ips) ==0:
                        outside_count+=1
                        print("outside_count=",outside_count)
                else:
                        inside_count += 1
                        print("inside_count=",inside_count)

        print("outside_count=",outside_count)
        print("inside_count=",inside_count)




def get_atom_counts_in_bin(run_number):

        import json
        with open('fe.json','r') as fe:
                fe_data = json.load(fe)
        fe.close()
        with open('mg.json','r') as mg:
                mg_data = json.load(mg)
        mg.close()
        with open('si.json','r') as si:
                si_data = json.load(si)
        si.close()
        with open('o.json','r') as o:
                o_data = json.load(o)
        o.close()

        buckets_fe = {}
        for i in range(NO_OF_BUCKET):
                buckets_fe[i] = 0

        for each in fe_data[run_number]:
                datum = each['atom_coordinate']
                x_cord, y_cord, z_cord = datum[0], datum[1], datum[2]
                x_offset = int((x_cord - LEFT_BOUNDARY)/BUCKET_LENGTH)
                y_offset = int((y_cord - LEFT_BOUNDARY)/BUCKET_LENGTH) 
                z_offset = int((z_cord - LEFT_BOUNDARY)/BUCKET_LENGTH)

                z_multiplier = round(NO_OF_BUCKET ** (2/3),2)
                y_multiplier = round(NO_OF_BUCKET ** (1/3),2)

                bucket_no = int((z_multiplier * z_offset + y_multiplier * y_offset + x_offset))
                try:
                    a = buckets_fe[bucket_no]
                    a = a + 1
                    buckets_fe[bucket_no] = a
                except Exception as e:
                    continue

        buckets_mg = {}
        for i in range(NO_OF_BUCKET):
                buckets_mg[i] = 0

        for each in mg_data[run_number]:
                datum = each['atom_coordinate']
                x_cord, y_cord, z_cord = datum[0], datum[1], datum[2]
                x_offset = int((x_cord - LEFT_BOUNDARY)/BUCKET_LENGTH)
                y_offset = int((y_cord - LEFT_BOUNDARY)/BUCKET_LENGTH) 
                z_offset = int((z_cord - LEFT_BOUNDARY)/BUCKET_LENGTH)

                z_multiplier = round(NO_OF_BUCKET ** (2/3),2)
                y_multiplier = round(NO_OF_BUCKET ** (1/3),2)

                bucket_no = int((z_multiplier * z_offset + y_multiplier * y_offset + x_offset))
                try:
                    a = buckets_mg[bucket_no]
                    a = a + 1
                    buckets_mg[bucket_no] = a
                except Exception as e:
                    continue

        buckets_si = {}
        for i in range(NO_OF_BUCKET):
                buckets_si[i] = 0

        for each in si_data[run_number]:
                datum = each['atom_coordinate']
                x_cord, y_cord, z_cord = datum[0], datum[1], datum[2]
                x_offset = int((x_cord - LEFT_BOUNDARY)/BUCKET_LENGTH)
                y_offset = int((y_cord - LEFT_BOUNDARY)/BUCKET_LENGTH) 
                z_offset = int((z_cord - LEFT_BOUNDARY)/BUCKET_LENGTH)

                z_multiplier = round(NO_OF_BUCKET ** (2/3),2)
                y_multiplier = round(NO_OF_BUCKET ** (1/3),2)

                bucket_no = int((z_multiplier * z_offset + y_multiplier * y_offset + x_offset))
                try:
                    a = buckets_si[bucket_no]
                    a = a + 1
                    buckets_si[bucket_no] = a
                except Exception as e:
                    continue

        buckets_o = {}
        for i in range(NO_OF_BUCKET):
                buckets_o[i] = 0

        for each in o_data[run_number]:
                datum = each['atom_coordinate']
                x_cord, y_cord, z_cord = datum[0], datum[1], datum[2]
                x_offset = int((x_cord - LEFT_BOUNDARY)/BUCKET_LENGTH)
                y_offset = int((y_cord - LEFT_BOUNDARY)/BUCKET_LENGTH) 
                z_offset = int((z_cord - LEFT_BOUNDARY)/BUCKET_LENGTH)

                z_multiplier = round(NO_OF_BUCKET ** (2/3),2)
                y_multiplier = round(NO_OF_BUCKET ** (1/3),2)

                bucket_no = int((z_multiplier * z_offset + y_multiplier * y_offset + x_offset))
                try:
                    a = buckets_o[bucket_no]
                    a = a + 1
                    buckets_o[bucket_no] = a
                except Exception as e:
                    continue

        with open("output_bin_atoms_{}.txt".format(run_number), "w") as file_object:
                for i in range(NO_OF_BUCKET):
                        line = "Bin Number -> {} : {}(Fe)|{}(Mg)|{}(Si)|{}(O) \n".format(i+1,buckets_fe[i],buckets_mg[i],buckets_si[i],buckets_o[i])
                        file_object.write(line)

        file_object.close()


def visualize_data_points(run_number):

        import json
        import pandas as pd
        import plotly.graph_objects as go
        import plotly.express as px
        with open('data/dp4000K/300_fe_last.json','r') as fe:
                try:
                        fe_json = json.load(fe)
                        if type(fe_json[0]) == dict:
                                raise Exception
                        fe_data = fe_json[run_number]
                        print("len-",len(fe_data))
                except Exception as e:
                        fe_data = fe_json
        fe.close()
        with open('data/dp4000K/300_mg_last.json','r') as mg:
                try:
                        mg_json = json.load(mg)
                        if type(mg_json[0]) == dict:
                                raise Exception
                        mg_data = mg_json[run_number]
                except Exception as e:
                        mg_data = mg_json
        mg.close()
        with open('data/dp4000K/300_si_last.json','r') as si:
                try:
                        si_json = json.load(si)
                        if type(si_json[0]) == dict:
                                raise Exception
                        si_data = si_json[run_number]
                except Exception as e:
                        si_data = si_json
        si.close()
        with open('data/dp4000K/300_o_last.json','r') as o:
                try:
                        o_json = json.load(o)
                        if type(o_json[0]) == dict:
                                raise Exception
                        o_data = o_json[run_number]
                except Exception as e:
                        o_data = o_json
        o.close()

        x_len = NO_OF_BUCKET**(1/3)
        def get_relative_coordinate(abs_coordinate,axis):
                #Alongside relative coordinate, ALSO take care of translation
                val = (abs_coordinate - LEFT_BOUNDARY)/(RIGHT_BOUNDARY - LEFT_BOUNDARY) * x_len
                mid_value = x_len/2
                if axis in [1,2]: # x and y axis apply boundary condition
                        if val > mid_value:
                                val = val - mid_value
                        else:
                                val = val + mid_value

                        copy_val = val
                        #translation
                        val = val + mid_value

                return round(val,2)

        print("fe-data=",fe_data)
        fe_coordinates_x = [get_relative_coordinate(obj['atom_coordinate'][0],1) for obj in fe_data]
        fe_coordinates_y = [get_relative_coordinate(obj['atom_coordinate'][1],2) for obj in fe_data]
        fe_coordinates_z = [get_relative_coordinate(obj['atom_coordinate'][2],3) for obj in fe_data]

        mg_coordinates_x = [get_relative_coordinate(obj['atom_coordinate'][0],1) for obj in mg_data]
        mg_coordinates_y = [get_relative_coordinate(obj['atom_coordinate'][1],2) for obj in mg_data]
        mg_coordinates_z = [get_relative_coordinate(obj['atom_coordinate'][2],3) for obj in mg_data]

        si_coordinates_x = [get_relative_coordinate(obj['atom_coordinate'][0],1) for obj in si_data]
        si_coordinates_y = [get_relative_coordinate(obj['atom_coordinate'][1],2) for obj in si_data]
        si_coordinates_z = [get_relative_coordinate(obj['atom_coordinate'][2],3) for obj in si_data]

        o_coordinates_x = [get_relative_coordinate(obj['atom_coordinate'][0],1) for obj in o_data]
        o_coordinates_y = [get_relative_coordinate(obj['atom_coordinate'][1],2) for obj in o_data]
        o_coordinates_z = [get_relative_coordinate(obj['atom_coordinate'][2],3) for obj in o_data]


        import math
        x_max = math.ceil(max(fe_coordinates_x))
        x_min = math.ceil(min(fe_coordinates_x))

        y_max = math.ceil(max(fe_coordinates_y))
        y_min = math.ceil(min(fe_coordinates_y))

        z_max = math.ceil(max(fe_coordinates_z))
        z_min = math.ceil(min(fe_coordinates_z))


        # DO NOT USE THIS!!!!! USE UPDATED : get_bin_no_updated()
        def get_bin_no(x_cords,y_cords,z_cords):
                list_bin_nos = []
                data = zip(x_cords,y_cords,z_cords)
                for (x_cord,y_cord,z_cord) in data:

                        # Ceil for x axis and int for others Logic!
                        x_offset = math.ceil(((x_cord - x_min)/(x_max-x_min)) * x_len)
                        y_offset = int(((y_cord - y_min)/(y_max-y_min)) * x_len)
                        z_offset = int(((z_cord - z_min)/(z_max-z_min)) * x_len)


                        z_multiplier = round(NO_OF_BUCKET ** (2/3),2)
                        y_multiplier = round(NO_OF_BUCKET ** (1/3),2)
                        bucket_no = int((z_multiplier * z_offset + y_multiplier * y_offset + x_offset))
                        list_bin_nos.append(bucket_no)
 
                return list_bin_nos


        fe_dict = {
                "x": fe_coordinates_x,
                "y": fe_coordinates_y,
                "z": fe_coordinates_z,
                "atom":["Fe"]*len(fe_coordinates_x),
                "bin_no" : get_bin_no(fe_coordinates_x,fe_coordinates_y,fe_coordinates_z)
        }

        mg_dict = {
                "x": mg_coordinates_x,
                "y": mg_coordinates_y,
                "z": mg_coordinates_z,
                "atom": ["Mg"]*len(mg_coordinates_x),
                "bin_no" : get_bin_no(mg_coordinates_x,mg_coordinates_y,mg_coordinates_z)
        }

        si_dict = {
                "x": si_coordinates_x,
                "y": si_coordinates_y,
                "z": si_coordinates_z,
                "atom": ["Si"]*len(si_coordinates_x),
                "bin_no" : get_bin_no(si_coordinates_x,si_coordinates_y,si_coordinates_z)
        }

        o_dict = {
                "x": o_coordinates_x,
                "y": o_coordinates_y,
                "z": o_coordinates_z,
                "atom":["O"]*len(o_coordinates_x),
                "bin_no": get_bin_no(o_coordinates_x,o_coordinates_y,o_coordinates_z)
        }

        df_fe = pd.DataFrame(fe_dict)
        df_mg = pd.DataFrame(mg_dict)
        df_si = pd.DataFrame(si_dict)
        df_o = pd.DataFrame(o_dict)

        frames = [df_fe, df_mg, df_si, df_o]

        df = pd.concat(frames)

        print("df_initial =",df.head())
        # conditions = [
        #       (df["z"] >=0) & (df["z"] <1),
        #       (df["z"] >=1) & (df["z"] <2),
        #       (df["z"] >=2) & (df["z"] <3),
        #       (df["z"] >=3) & (df["z"] <4),
        #       (df["z"] >=4) & (df["z"] <5),
        #       (df["z"] >=5) & (df["z"] <6),
        #       (df["z"] >=6) & (df["z"] <7),
        #       (df["z"] >=7) & (df["z"] <8),
        #       (df["z"] >=8) & (df["z"] <9),
        #       (df["z"] >=9) & (df["z"] <10),
        #       (df["z"] >=10) & (df["z"] <11),
        #       (df["z"] >=11) & (df["z"] <12),
        #       (df["z"] >=12) & (df["z"] <13),
        #       (df["z"] >=13) & (df["z"] <14),
        #       (df["z"] >=14) & (df["z"] <15),
        #       (df["z"] >=15) & (df["z"] <16)
        # ]

        # values = ["1","2","3","4","5","6","7","8","9","10","11","12","13","14","15","16"]
        # x_len = int(round(x_len,2))
        # conditions = conditions[:x_len]
        # values = values[:x_len]
        # import numpy as np
        # df['layer'] = np.select(conditions,values)

        # def translate_df(df):
        #       df.loc[:,'x'] += 2
        #       return df

        # df = translate_df(df)

        x_min = min(df['x'])
        y_min = min(df['y'])
        z_min = min(df['z'])

        df['x_trans'] = df.apply(lambda x : x['x'] - x_min,axis=1)
        df['y_trans'] = df.apply(lambda x : x['y'] - y_min,axis=1)
        df['z_trans'] = df.apply(lambda x : x['z'] - z_min,axis=1)

        conditions = [
                (df["z_trans"] >=0) & (df["z_trans"] <1),
                (df["z_trans"] >=1) & (df["z_trans"] <2),
                (df["z_trans"] >=2) & (df["z_trans"] <3),
                (df["z_trans"] >=3) & (df["z_trans"] <4),
                (df["z_trans"] >=4) & (df["z_trans"] <5),
                (df["z_trans"] >=5) & (df["z_trans"] <6),
                (df["z_trans"] >=6) & (df["z_trans"] <7),
                (df["z_trans"] >=7) & (df["z_trans"] <8),
                (df["z_trans"] >=8) & (df["z_trans"] <9),
                (df["z_trans"] >=9) & (df["z_trans"] <10),
                (df["z_trans"] >=10) & (df["z_trans"] <11),
                (df["z_trans"] >=11) & (df["z_trans"] <12),
                (df["z_trans"] >=12) & (df["z_trans"] <13),
                (df["z_trans"] >=13) & (df["z_trans"] <14),
                (df["z_trans"] >=14) & (df["z_trans"] <15),
                (df["z_trans"] >=15) & (df["z_trans"] <16)
        ]

        values = ["1","2","3","4","5","6","7","8","9","10","11","12","13","14","15","16"]
        x_len = int(round(x_len,2))
        conditions = conditions[:x_len]
        values = values[:x_len]
        import numpy as np
        df['layer'] = np.select(conditions,values)

        def get_bin_no_updated(x):
                z_multiplier = round(NO_OF_BUCKET ** (2/3),2)
                y_multiplier = round(NO_OF_BUCKET ** (1/3),2)
                import math
                x_offset = math.ceil(x['x_trans'])
                y_offset = int(x['y_trans'])
                z_offset = int(x['z_trans'])

                bucket_no = int((z_multiplier * z_offset + y_multiplier * y_offset + x_offset))

                return bucket_no


        df['bin_no_updated'] = df.apply(lambda x: get_bin_no_updated(x),axis=1)
# dataframe.loc[dataframe['Percentage'] > 70] 
        filtered_df = df[df['atom'] == 'Fe']
        df['fe_count_in_bin'] = df.apply(lambda x: filtered_df['bin_no_updated'].value_counts().get(x['bin_no_updated'],0),axis=1)
        # df['fe_count_in_bin'] = df.apply(lambda x: df.loc[df['atom'] ==  'Fe']['bin_no'].value_counts()[x['bin_no']],axis=1)


        #plot all 8 grids .... x_len is the number of boxes in each axis
        print('shape1=',df.shape)
        #All filtering happens here
        filter_val = args.filter
        filter_value_fe_count_max = int(filter_val.split('-')[2])
        filter_value_fe_count_min = int(filter_val.split('-')[1])
        filter_atoms = filter_val.split('-')[0]
        filter_atoms = filter_atoms.split(',')
        df = df[df['atom'].isin(filter_atoms)]
        # if filter_atom == "fe":
        #       df = df[df['atom']=='Fe']
        # elif filter_atom == "mg":
        #       df = df[df['atom']=='Mg']
        # elif filter_atom == "si":
        #       df = df[df['atom']=='Si']
        # elif filter_atom == "o":
        #       df = df[df['atom'] == 'O']
        # else:
        #       pass


        final_df = df[(df['fe_count_in_bin']>= filter_value_fe_count_min) & (df['fe_count_in_bin']<=filter_value_fe_count_max)]
        print("shape2=",final_df.shape)

        def boundary_condition_in_df(df):
                # df['x_trans'] = df['x_trans'].apply(lambda x: x+4 if x<4 else x-4)
                # df['y_trans'] = df['y_trans'].apply(lambda y: y+4 if y<4 else y-4)

                df['x_trans'] = df['x_trans']
                df['y_trans'] = df['y_trans']

                return df

        final_df = boundary_condition_in_df(final_df)


        for i in range(x_len):
                df_filtered = final_df[final_df['layer']==str(i+1)] 
                print("i = {}, len df filter = {}".format(i+1,len(df_filtered)))
                fig = px.scatter(
                                                 df_filtered, 
                                                 x="x_trans",
                                                 y="y_trans",
                                                 color="atom",
                                                 height=700,
                                                 width=800)
                fig.update_layout(
                                            margin=dict(l=20, r=20, t=20, b=20),
                                            paper_bgcolor="LightSteelBlue")
                fig.update_traces(
                marker_size=5)

                fig = go.Figure(fig,layout_xaxis_range=[0,x_len],layout_yaxis_range=[0,x_len])
                fig.write_image("Layer_{}.png".format(i+1))


        combine_images(columns=4, space=20, images=['Layer_1.png', 'Layer_2.png', 'Layer_3.png', 'Layer_4.png', 'Layer_5.png', 'Layer_6.png', 'Layer_7.png','Layer_8.png'])


def visualize_data_points_individual(run_number):

        import json
        import pandas as pd
        import plotly.graph_objects as go
        import plotly.express as px
        with open('fe_last.json','r') as fe:
                try:
                        fe_json = json.load(fe)
                        if type(fe_json[0]) == dict:
                                raise Exception
                        fe_data = fe_json[run_number]
                        print("len-",len(fe_data))
                except Exception as e:
                        fe_data = fe_json
        fe.close()
        with open('mg_last.json','r') as mg:
                try:
                        mg_json = json.load(mg)
                        if type(mg_json[0]) == dict:
                                raise Exception
                        mg_data = mg_json[run_number]
                except Exception as e:
                        mg_data = mg_json
        mg.close()
        with open('si_last.json','r') as si:
                try:
                        si_json = json.load(si)
                        if type(si_json[0]) == dict:
                                raise Exception
                        si_data = si_json[run_number]
                except Exception as e:
                        si_data = si_json
        si.close()
        with open('o_last.json','r') as o:
                try:
                        o_json = json.load(o)
                        if type(o_json[0]) == dict:
                                raise Exception
                        o_data = o_json[run_number]
                except Exception as e:
                        o_data = o_json
        o.close()

        x_len = NO_OF_BUCKET**(1/3)
        def get_relative_coordinate(abs_coordinate,axis):
                #Alongside relative coordinate, ALSO take care of translation
                print(abs_coordinate,LEFT_BOUNDARY,RIGHT_BOUNDARY,x_len)
                val = (abs_coordinate - LEFT_BOUNDARY)/(RIGHT_BOUNDARY - LEFT_BOUNDARY) * x_len
                mid_value = x_len/2
                if axis in [1,2]: # x and y axis apply boundary condition
                        if val > mid_value:
                                val = val - mid_value
                        else:
                                val = val + mid_value

                        copy_val = val
                        #translation
                        val = val + mid_value

                return round(val,2)

        print("fe-data=",fe_data)
        fe_coordinates_x = [get_relative_coordinate(obj['atom_coordinate'][0],1) for obj in fe_data]
        fe_coordinates_y = [get_relative_coordinate(obj['atom_coordinate'][1],2) for obj in fe_data]
        fe_coordinates_z = [get_relative_coordinate(obj['atom_coordinate'][2],3) for obj in fe_data]

        mg_coordinates_x = [get_relative_coordinate(obj['atom_coordinate'][0],1) for obj in mg_data]
        mg_coordinates_y = [get_relative_coordinate(obj['atom_coordinate'][1],2) for obj in mg_data]
        mg_coordinates_z = [get_relative_coordinate(obj['atom_coordinate'][2],3) for obj in mg_data]

        si_coordinates_x = [get_relative_coordinate(obj['atom_coordinate'][0],1) for obj in si_data]
        si_coordinates_y = [get_relative_coordinate(obj['atom_coordinate'][1],2) for obj in si_data]
        si_coordinates_z = [get_relative_coordinate(obj['atom_coordinate'][2],3) for obj in si_data]

        o_coordinates_x = [get_relative_coordinate(obj['atom_coordinate'][0],1) for obj in o_data]
        o_coordinates_y = [get_relative_coordinate(obj['atom_coordinate'][1],2) for obj in o_data]
        o_coordinates_z = [get_relative_coordinate(obj['atom_coordinate'][2],3) for obj in o_data]


        print("\n")
        print("fe_coordinates_x=",fe_coordinates_x)

        import math
        x_max = math.ceil(max(fe_coordinates_x))
        x_min = math.ceil(min(fe_coordinates_x))

        y_max = math.ceil(max(fe_coordinates_y))
        y_min = math.ceil(min(fe_coordinates_y))

        z_max = math.ceil(max(fe_coordinates_z))
        z_min = math.ceil(min(fe_coordinates_z))

        print(x_max,x_min)


        # DO NOT USE THIS!!!!! USE UPDATED : get_bin_no_updated()
        def get_bin_no(x_cords,y_cords,z_cords):
                list_bin_nos = []
                data = zip(x_cords,y_cords,z_cords)
                for (x_cord,y_cord,z_cord) in data:

                        # Ceil for x axis and int for others Logic!
                        x_offset = math.ceil(((x_cord - x_min)/(x_max-x_min)) * x_len)
                        y_offset = int(((y_cord - y_min)/(y_max-y_min)) * x_len)
                        z_offset = int(((z_cord - z_min)/(z_max-z_min)) * x_len)


                        z_multiplier = round(NO_OF_BUCKET ** (2/3),2)
                        y_multiplier = round(NO_OF_BUCKET ** (1/3),2)
                        bucket_no = int((z_multiplier * z_offset + y_multiplier * y_offset + x_offset))
                        list_bin_nos.append(bucket_no)
 
                return list_bin_nos


        fe_dict = {
                "x": fe_coordinates_x,
                "y": fe_coordinates_y,
                "z": fe_coordinates_z,
                "atom":["Fe"]*len(fe_coordinates_x),
                "bin_no" : get_bin_no(fe_coordinates_x,fe_coordinates_y,fe_coordinates_z)
        }

        mg_dict = {
                "x": mg_coordinates_x,
                "y": mg_coordinates_y,
                "z": mg_coordinates_z,
                "atom": ["Mg"]*len(mg_coordinates_x),
                "bin_no" : get_bin_no(mg_coordinates_x,mg_coordinates_y,mg_coordinates_z)
        }

        si_dict = {
                "x": si_coordinates_x,
                "y": si_coordinates_y,
                "z": si_coordinates_z,
                "atom": ["Si"]*len(si_coordinates_x),
                "bin_no" : get_bin_no(si_coordinates_x,si_coordinates_y,si_coordinates_z)
        }

        o_dict = {
                "x": o_coordinates_x,
                "y": o_coordinates_y,
                "z": o_coordinates_z,
                "atom":["O"]*len(o_coordinates_x),
                "bin_no": get_bin_no(o_coordinates_x,o_coordinates_y,o_coordinates_z)
        }

        df_fe = pd.DataFrame(fe_dict)
        df_mg = pd.DataFrame(mg_dict)
        df_si = pd.DataFrame(si_dict)
        df_o = pd.DataFrame(o_dict)

        frames = [df_fe, df_mg, df_si, df_o]

        df = pd.concat(frames)

        print("df_initial =",df.head())
        # conditions = [
        #       (df["z"] >=0) & (df["z"] <1),
        #       (df["z"] >=1) & (df["z"] <2),
        #       (df["z"] >=2) & (df["z"] <3),
        #       (df["z"] >=3) & (df["z"] <4),
        #       (df["z"] >=4) & (df["z"] <5),
        #       (df["z"] >=5) & (df["z"] <6),
        #       (df["z"] >=6) & (df["z"] <7),
        #       (df["z"] >=7) & (df["z"] <8),
        #       (df["z"] >=8) & (df["z"] <9),
        #       (df["z"] >=9) & (df["z"] <10),
        #       (df["z"] >=10) & (df["z"] <11),
        #       (df["z"] >=11) & (df["z"] <12),
        #       (df["z"] >=12) & (df["z"] <13),
        #       (df["z"] >=13) & (df["z"] <14),
        #       (df["z"] >=14) & (df["z"] <15),
        #       (df["z"] >=15) & (df["z"] <16)
        # ]

        # values = ["1","2","3","4","5","6","7","8","9","10","11","12","13","14","15","16"]
        # x_len = int(round(x_len,2))
        # conditions = conditions[:x_len]
        # values = values[:x_len]
        # import numpy as np
        # df['layer'] = np.select(conditions,values)

        # def translate_df(df):
        #       df.loc[:,'x'] += 2
        #       return df

        # df = translate_df(df)

        x_min = min(df['x'])
        y_min = min(df['y'])
        z_min = min(df['z'])

        df['x_trans'] = df.apply(lambda x : x['x'] - x_min,axis=1)
        df['y_trans'] = df.apply(lambda x : x['y'] - y_min,axis=1)
        df['z_trans'] = df.apply(lambda x : x['z'] - z_min,axis=1)

        conditions = [
                (df["z_trans"] >=0) & (df["z_trans"] <1),
                (df["z_trans"] >=1) & (df["z_trans"] <2),
                (df["z_trans"] >=2) & (df["z_trans"] <3),
                (df["z_trans"] >=3) & (df["z_trans"] <4),
                (df["z_trans"] >=4) & (df["z_trans"] <5),
                (df["z_trans"] >=5) & (df["z_trans"] <6),
                (df["z_trans"] >=6) & (df["z_trans"] <7),
                (df["z_trans"] >=7) & (df["z_trans"] <8),
                (df["z_trans"] >=8) & (df["z_trans"] <9),
                (df["z_trans"] >=9) & (df["z_trans"] <10),
                (df["z_trans"] >=10) & (df["z_trans"] <11),
                (df["z_trans"] >=11) & (df["z_trans"] <12),
                (df["z_trans"] >=12) & (df["z_trans"] <13),
                (df["z_trans"] >=13) & (df["z_trans"] <14),
                (df["z_trans"] >=14) & (df["z_trans"] <15),
                (df["z_trans"] >=15) & (df["z_trans"] <16)
        ]

        values = ["1","2","3","4","5","6","7","8","9","10","11","12","13","14","15","16"]
        x_len = int(round(x_len,2))
        conditions = conditions[:x_len]
        values = values[:x_len]
        import numpy as np
        df['layer'] = np.select(conditions,values)

        def get_bin_no_updated(x):
                z_multiplier = round(NO_OF_BUCKET ** (2/3),2)
                y_multiplier = round(NO_OF_BUCKET ** (1/3),2)
                import math
                x_offset = math.ceil(x['x_trans'])
                y_offset = int(x['y_trans'])
                z_offset = int(x['z_trans'])

                bucket_no = int((z_multiplier * z_offset + y_multiplier * y_offset + x_offset))

                return bucket_no


        df['bin_no_updated'] = df.apply(lambda x: get_bin_no_updated(x),axis=1)
# dataframe.loc[dataframe['Percentage'] > 70] 
        filtered_df = df[df['atom'] == 'Fe']
        df['fe_count_in_bin'] = df.apply(lambda x: filtered_df['bin_no_updated'].value_counts().get(x['bin_no_updated'],0),axis=1)
        # df['fe_count_in_bin'] = df.apply(lambda x: df.loc[df['atom'] ==  'Fe']['bin_no'].value_counts()[x['bin_no']],axis=1)


        #plot all 8 grids .... x_len is the number of boxes in each axis
        print('shape1=',df.shape)
        #All filtering happens here
        filter_val = args.filter
        filter_value_fe_count_max = int(filter_val.split('-')[2])
        filter_value_fe_count_min = int(filter_val.split('-')[1])
        filter_atoms = filter_val.split('-')[0]
        filter_atoms = filter_atoms.split(',')
        df = df[df['atom'].isin(filter_atoms)]
        # if filter_atom == "fe":
        #       df = df[df['atom']=='Fe']
        # elif filter_atom == "mg":
        #       df = df[df['atom']=='Mg']
        # elif filter_atom == "si":
        #       df = df[df['atom']=='Si']
        # elif filter_atom == "o":
        #       df = df[df['atom'] == 'O']
        # else:
        #       pass

        # 2nd filter as per bijaya sir
        filter_val_2 = args.filter2
        filter_value_fe_count_max_2 = int(filter_val_2.split('-')[2])
        filter_value_fe_count_min_2 = int(filter_val_2.split('-')[1])
        filter_atoms_2 = filter_val_2.split('-')[0]
        filter_atoms_2 = filter_atoms_2.split(',')
        df_2 = df[df['atom'].isin(filter_atoms_2)]

        # 3rd filter as per bijaya sir (for intermediate region)
        filter_val_3 = args.filter3
        filter_value_fe_count_max_3 = int(filter_val_3.split('-')[2])
        filter_value_fe_count_min_3= int(filter_val_3.split('-')[1])
        filter_atoms_3 = filter_val_3.split('-')[0]
        filter_atoms_3 = filter_atoms_3.split(',')
        df_3 = df[df['atom'].isin(filter_atoms_3)]


        final_df = df[(df['fe_count_in_bin']>= filter_value_fe_count_min) & (df['fe_count_in_bin']<=filter_value_fe_count_max)]
        final_df['filter_type'] = 'lower'

        final_df_2 = df_2[(df_2['fe_count_in_bin']>= filter_value_fe_count_min_2) & (df_2['fe_count_in_bin']<=filter_value_fe_count_max_2)]
        final_df_2['filter_type'] = 'upper'


        final_df_3 = df_3[(df_3['fe_count_in_bin']>= filter_value_fe_count_min_3) & (df_3['fe_count_in_bin']<=filter_value_fe_count_max_3)]
        final_df_3['filter_type'] = 'middle'


        print("--------")
        print(final_df.shape)
        print(final_df_2.shape)
        print(filter_value_fe_count_min)
        print(filter_value_fe_count_min_2)

        def boundary_condition_in_df(df):

                # For dpmd
                # df['x_trans'] = df['x_trans'].apply(lambda x: x+4 if x<4 else x-4)
                # df['y_trans'] = df['y_trans'].apply(lambda y: y+4 if y<4 else y-4)

                #For fpmd
                df['x_trans'] = df['x_trans'].apply(lambda x: x+2 if x<2 else x-2)
                df['y_trans'] = df['y_trans'].apply(lambda y: y+2 if y<2 else y-2)              

                # df['x_trans'] = df['x_trans']
                # df['y_trans'] = df['y_trans']

                return df

        #combine lower and upper
        combine_df = [final_df,final_df_2,final_df_3]
        finalest_df = pd.concat(combine_df)


        finalest_df = boundary_condition_in_df(finalest_df)

        print("finalest_df-",finalest_df.head(20))
        print("x_len-",x_len)


        images = []

        for i in range(x_len):

                images.append("Layer_{}.png".format(i+1))
                df_filtered = finalest_df[finalest_df['layer']==str(i+1)] 
                print("i = {}, len df filter = {}".format(i+1,len(df_filtered)))
                fig = px.scatter(
                                                 df_filtered, 
                                                 x="x_trans",
                                                 y="y_trans",
                                                 color="filter_type",
                                                 height=700 if x_len>4 else 350,
                                                 width=800 if x_len>4 else 400,
                                                 color_discrete_map = {'upper' : 'red','lower' : 'blue', 'middle' : 'green'}
                                                 )
                fig.update_layout(
                                            margin=dict(l=20, r=20, t=20, b=20),
                                            paper_bgcolor="LightSteelBlue")
                fig.update_traces(
                marker_size=5)

                fig = go.Figure(fig,layout_xaxis_range=[0,x_len],layout_yaxis_range=[0,x_len])
                fig.write_image("Layer_{}.png".format(i+1))


        # combine_images(columns=4, space=20, images=['Layer_1.png', 'Layer_2.png', 'Layer_3.png', 'Layer_4.png', 'Layer_5.png', 'Layer_6.png', 'Layer_7.png','Layer_8.png'])
        combine_images(columns=4, space=20, images=images)

def combine_images(columns, space, images):
        from PIL import Image
        rows = len(images) // columns
        if len(images) % columns:
            rows += 1
        width_max = max([Image.open(image).width for image in images])
        height_max = max([Image.open(image).height for image in images])
        background_width = width_max*columns + (space*columns)-space
        background_height = height_max*rows + (space*rows)-space
        background = Image.new('RGBA', (background_width, background_height), (255, 255, 255, 255))
        x = 0
        y = 0
        for i, image in enumerate(images):
            img = Image.open(image)
            x_offset = int((width_max-img.width)/2)
            y_offset = int((height_max-img.height)/2)
            background.paste(img, (x+x_offset, y+y_offset))
            x += width_max + space
            if (i+1) % columns == 0:
                y += height_max + space
                x = 0
        background.save('image.png')


def parsed_data_dpmd(data):
        x_data = float(data[2])
        y_data = float(data[3])
        z_data = float(data[4])

        #x_cord
        # if x_data > CELL_SIZE/2:
        #     x_data = x_data - CELL_SIZE

        # #y_cord
        # if y_data > CELL_SIZE/2:
        #     y_data = y_data - CELL_SIZE

        # #z_cord
        # if z_data > CELL_SIZE/2:
        #     z_data = z_data - CELL_SIZE

        return {
            "atom_id" : data[0],
            "atom_class" : data[1],
            "atom_coordinate" : [x_data,y_data,z_data] 
        }

def get_points_dpmd(path,num):
        file = open(path,"r")
        read_it = file.read()

        lines = read_it.splitlines()
        check_next = False
        initial = True
        flag = 0
        final_data = []
        one_snap = []

        for line in lines:
                if check_next:
                        NUMBER_OF_ATOMS = int(line.strip())
                        check_next = False
                        break

                if line == "ITEM: NUMBER OF ATOMS" and initial:
                        initial = False
                        check_next = True

        for i,line in enumerate(lines):
                line = line.strip()
                if "ITEM: ATOMS id type x y z" in line:
                        if flag:
                                final_data.append(one_snap)
                                one_snap = []
                        flag = 1
                match = re.match("([0-9]*)([ ]*)([0-9]*)([ ]*)([0-9]*\.*\d*)([ ]*)([0-9]*\.*\d*)([ ]*)([0-9]*\.*\d*)",line)
                if match:
                        if match.group(3) == num:

                                one = match.group(1)
                                two = match.group(3)
                                three = match.group(5)
                                four = match.group(7)
                                five = match.group(9)

                                val = parsed_data_dpmd([one if one else 0,two if two else 0,three if three else 0,four if four else 0,five if five else 0])
                                one_snap.append(val)

        pts = [[a for a in sorted(individual_data,key=lambda x:int(x['atom_id']))] for individual_data in final_data]
        return pts

def parse_dpmd_data(file_to_parse):

        for i in ['1','2','3','4','5']:
                data = get_points_dpmd(file_to_parse,i)

                # Get last snapshot
                last_data = len(data) -1 
                import json
                if i=="1":
                        out_file = open("fe.json", "w")
                        out_file_last = open("fe_last.json","w")
                elif i=="2":
                        out_file = open("mg.json", "w")
                        out_file_last = open("mg_last.json","w")
                elif i=="3":
                        out_file = open("si.json","w")
                        out_file_last = open("si_last.json","w")
                elif i=="4":
                        out_file = open("o.json","w")
                        out_file_last = open("o_last.json","w")
                else:
                        out_file = open("n.json","w")
                        out_file_last = open("n_last.json","w")

                json.dump(data, out_file, indent = 6)
                json.dump(data[last_data], out_file_last, indent = 6)
                  
                out_file.close()

def merge_analysis():
        from pypdf import PdfMerger

        import os 
          
        # Get the list of all files and directories 
        # in the root directory 
        path = "."
        dir_list = os.listdir(path)

        print(dir_list)
        pdfs = [dir for dir in dir_list if "AnalysisNumber_" in dir]
        pdfs.sort()
        if not pdfs:
                print("No AnalysisNumber_ files")
                exit(1)

        merger = PdfMerger()

        for pdf in pdfs:
            merger.append(pdf)

        merger.write("Analysis.pdf")
        merger.close()

if __name__ == '__main__':
        parser = argparse.ArgumentParser(description='Parser')
        parser.add_argument('-n','--no_of_atoms', help='No of atoms',required=True)
        parser.add_argument('-s','--cell_size', help='Maximimum cell size',required=True)
        parser.add_argument('-b','--no_of_buckets', help='No of buckets',required=True)
        parser.add_argument('-p','--percent', help='nth simulation',required=True)
        parser.add_argument('-i','--input', help='Input Json File',required=False)
        parser.add_argument('-j','--input_subatom', help='Input Subatom Json File',required=False)
        parser.add_argument('-o','--output', help='Output File Name',required=True)
        parser.add_argument('-k','--to_do', help='What to perform',required=True)
        parser.add_argument('-f','--filter', help='Filter Values',required=False)
        parser.add_argument('-g','--filter2', help='Filter 2nd Values',required=False)
        parser.add_argument('-q','--filter3',help='Filter Intermediate Values',required=False)
        parser.add_argument('-m','--merge', help='What data to merge',required=False)
        parser.add_argument('-d','--sub_distribution', help='Sub Distribution',required=False)
        parser.add_argument('-a','--path',help='Path of clusterd_within_fe.json',required=False)
        # parser.add_argument('-c','--convert',help='Convert XDATCAR to JSON',required=False)
        #name of atom for which the simulation path is being generated
        parser.add_argument('-z','--atom_for_sim_path',help='Which atom for simulation path?',required=False)
        parser.add_argument('-e','--boundary',help='LEFT BOUNDARY 0 or -ve',required=False)
        parser.add_argument('-r','--run_number',help='LEFT BOUNDARY 0 or -ve',required=False)
        parser.add_argument('-l','--dpmd_file',help='dpmd file',required=False)





        args = parser.parse_args()


        NO_OF_ATOMS = int(args.no_of_atoms)
        CELL_SIZE = int(args.cell_size)
        NO_OF_BUCKET = int(args.no_of_buckets)

        BOUNDARY = (round(-1 * CELL_SIZE/2,2),round(CELL_SIZE/2,2))
        # This is the usual one
        print(args.boundary)
        try:
                if args.boundary == "0" or args.boundary is None:
                        LEFT_BOUNDARY = 0
                        RIGHT_BOUNDARY = CELL_SIZE #Hard code for now
                else:
                        print("a")
                        LEFT_BOUNDARY = BOUNDARY[0]
                        RIGHT_BOUNDARY = BOUNDARY[1]
        except Exception as e:
                print("exception =",e)
                LEFT_BOUNDARY = 0
                RIGHT_BOUNDARY = CELL_SIZE

        print("LEFT_BOUNDARY=",LEFT_BOUNDARY)

        BUCKET_LENGTH = CELL_SIZE/ (NO_OF_BUCKET**(1/3)) + 0.000001 #adding val for boundary conditions

        AVG_ATOMS_PER_BUCKET = int(NO_OF_ATOMS/NO_OF_BUCKET)
        #setting max atom a bucket can have for x-axis
        print("cell size =",CELL_SIZE)
        print("number of atoms=",NO_OF_ATOMS)
        print("number of bucket =",NO_OF_BUCKET)
        MAX_ATOMS_PER_BUCKET = 2 * AVG_ATOMS_PER_BUCKET
        print("MAX_ATOMS_PER_BUCKET=",MAX_ATOMS_PER_BUCKET)
        if args.input:
                all_data = get_points(args.input)
        else:
                all_data = []
        #%percentile of data
        n_data =float(args.percent)

        if n_data > 1 or n_data <0:
                val = len(all_data) - 1
        else:
                #Percentage of nth data
                val = int(n_data * len(all_data)) - 1
                # val = 455

        if args.to_do == "distribution":
                get_atom_distribution(all_data[val],graph_output=args.output)
        elif args.to_do == "filter":
                f = args.filter
                if not f:
                        print("Filter Parameter not passed")
                        exit(1)
                start = f.split('-')[0]
                end = f.split('-')[1]
                get_filtered_data(all_data[val],start,end)
        elif args.to_do == "merge":
                merge_val = args.merge.split("-")
                merge_data(merge_val)
        elif args.to_do == "nested_cluster":
                f = args.filter
                if not f:
                        print("Filter Parameter not passed")
                        exit(1)
                start = f.split('-')[0]
                end = f.split('-')[1]
                atom_class = f.split('-')[2]
                sub_distribution = args.sub_distribution
                print("sub-",sub_distribution)
                get_filtered_data_all(all_data[val],start,end,atom_class,sub_distribution)
        elif args.to_do == "proximity":
                path = args.path 
                with open(path, 'r') as myfile:
                        data = json.load(myfile)
                get_proximity(data)
        elif args.to_do == "correlation_test":
                scatterplot()
        elif args.to_do == 'tsne':
                tsne(all_data[val])
        elif args.to_do == 'atom_movement':
                atom_class = int(args.atom_for_sim_path)
                atom_movement_data(all_data,str(atom_class))
        elif args.to_do == "xdatcar_to_json":
                print("-----------------------------")
                path = args.path
                print(path)
                xdatcar_to_json(all_data,path)
        elif args.to_do == "average_distribution":
                get_atom_distribution_average(all_data)
        elif args.to_do == "get_sub_atom_count_in_fe":
                all_data_n = get_points(args.input_subatom)
                get_sub_atom_count_in_fe(all_data,all_data_n)
        elif args.to_do == "connected_components":
                f = args.filter
                start = f.split('-')[0]
                end = f.split('-')[1]
                get_connected_componen(all_data[val],start,end)
        elif args.to_do == "alphashapes":
                import json
                path = args.path

                i_data = args.input
                with open(i_data,'r') as ip:
                        input_data = json.load(ip)

                with open(path,'r') as f:
                        data = json.load(f)
                
                #data is for building alpha shape, input_data are the points that need to be checked if they are within polygon
                atom_count_in_alpha_shapes(data,input_data,n_data)


        elif args.to_do == "see_bin_data":
                try:
                        run_number = int(args.run_number)
                except TypeError:
                        run_number = val
                        print("run_number=",run_number)
                get_atom_counts_in_bin(run_number)
        elif args.to_do == "visualize_data_points":
                try:
                        run_number = int(args.run_number)
                except TypeError:
                        run_number = val
                        print("run_number=",run_number)

                visualize_data_points(run_number)

        elif args.to_do == "visualize_data_points_individual":
                try:
                        run_number = int(args.run_number)
                except TypeError:
                        run_number = val
                        print("run_number=",run_number)

                visualize_data_points_individual(run_number)

        elif args.to_do == "parse_dpmd_data":
                file_to_parse = args.dpmd_file
                parse_dpmd_data(file_to_parse)
        elif args.to_do == "merge_analysis":
                merge_analysis()

        else:
                exit(1)





# Bin plot in 3d space
# Spatial with VOXELS.py

# import re


# import sys
# import argparse
# parser = argparse.ArgumentParser(description='Parser')
# parser.add_argument('-n','--no_of_atoms', help='No of atoms',required=True)
# parser.add_argument('-s','--cell_size', help='Maximimum cell size',required=True)
# parser.add_argument('-b','--no_of_buckets', help='No of buckets',required=True)
# parser.add_argument('-o','--output', help='output_file_name',required=True)
# parser.add_argument('-i','--input', help='Input Json File',required=True)
# args = parser.parse_args()

# NO_OF_ATOMS = int(args.no_of_atoms)
# CELL_SIZE = int(args.cell_size)
# NO_OF_BUCKET = int(args.no_of_buckets)

# BOUNDARY = (round(-1 * CELL_SIZE/2,2),round(CELL_SIZE/2,2))
# LEFT_BOUNDARY = BOUNDARY[0]
# RIGHT_BOUNDARY = BOUNDARY[1]

# BUCKET_LENGTH = CELL_SIZE/ (NO_OF_BUCKET**(1/3)) + 0.1 #adding val for boundary conditions

# AVG_ATOMS_PER_BUCKET = int(NO_OF_ATOMS/NO_OF_BUCKET)
# #setting max atom a bucket can have for x-axis
# MAX_ATOMS_PER_BUCKET = 2 * AVG_ATOMS_PER_BUCKET

# def parsed_data(data):

#     x_data = float(data[2])
#     y_data = float(data[3])
#     z_data = float(data[4])

#     #x_cord
#     if x_data > CELL_SIZE/2:
#         x_data = x_data - CELL_SIZE

#     #y_cord
#     if y_data > CELL_SIZE/2:
#         y_data = y_data - CELL_SIZE

#     #z_cord
#     if z_data > CELL_SIZE/2:
#         z_data = z_data - CELL_SIZE



#     return {
#         "atom_id" : data[0],
#         "atom_class" : data[1],
#         "atom_coordinate" : [x_data,y_data,z_data] 
#     }

# def get_points(path):
#     import json
#     with open(path, 'r') as myfile:
#         final_data = json.load(myfile)

#     pts = [[pt['atom_coordinate'] for pt in pts] for pts in final_data]
#     return pts

# all_data = get_points(args.input)
# buckets = {}
# no_of_samples = 0

# def countX(lst, x):
#     count = 0
#     for ele in lst:
#         if (ele == x):
#             count = count + 1
#     return count

# my_distro = {}
# for i in range(MAX_ATOMS_PER_BUCKET+1):
#     my_distro[i] = 0

# #Only using 1 data
# all_data = [all_data[0]]

# for i,each_data in enumerate(all_data):
#     if not 1:
#         continue
#     else:
#         no_of_samples = no_of_samples + 1
#     for i in range(NO_OF_BUCKET):
#         buckets[i] = 0
#     for datum in each_data:
#         x_cord, y_cord, z_cord = datum[0], datum[1], datum[2]
#         x_offset = int((x_cord - LEFT_BOUNDARY)/BUCKET_LENGTH)
#         y_offset = int((y_cord - LEFT_BOUNDARY)/BUCKET_LENGTH) 
#         z_offset = int((z_cord - LEFT_BOUNDARY)/BUCKET_LENGTH)

#         z_multiplier = round(NO_OF_BUCKET ** (2/3),2)
#         y_multiplier = round(NO_OF_BUCKET ** (1/3),2)

#         bucket_no = int((z_multiplier * z_offset + y_multiplier * y_offset + x_offset))
#         try:
#             a = buckets[bucket_no]
#             a = a + 1
#             buckets[bucket_no] = a
#         except Exception as e:
#             continue

#     list_num_per_bucket = []
#     """later"""
#     return_buckets = []
#     start = 10
#     end=30
#     for x,y in buckets.items():
#         list_num_per_bucket.append(y)
#         if y >=start and y<=end:
#             return_buckets.append(x)

#     unique_elements = list(set(list_num_per_bucket))

#     for element in unique_elements:
#         count = countX(list_num_per_bucket,element)
#         my_distro[element] = my_distro[element] + count



# return_buckets_coordiate = []
# k = NO_OF_BUCKET ** (1/3)
# for num in return_buckets:
#     x = int((num - 1)/k**2) + 1
#     y = int(((num - (x-1)*k**2) - 1)/k) + 1
#     z = (num-1) - (k**2*(x-1)) - (k*(y-1)) + 1
#     return_buckets_coordiate.append([round(x,2),round(y,2),round(z,2)])


# print(return_buckets_coordiate)

# final_frequency = {}
# for key,value in my_distro.items():
#     final_frequency[key] = round(value / no_of_samples,2)

# data = []
# for x,y in final_frequency.items():
#     data = data + int(round(y,2)) * [x]


# import matplotlib.pyplot as plt
# import numpy as np
# import matplotlib
# #matplotlib.use('TKAgg')
# num_bins = 50

# final_list = []
# for i in data:
#     final_list.append(int(round(i,2)))

# set_list = set(final_list)
# length = len(set_list)

# from datetime import datetime
# now = datetime.now()
# timestamp = int(now.timestamp())


# try:
#     n, bins, patches = plt.hist(final_list, length, facecolor='blue', alpha=0.5,edgecolor="red",align='mid' )
#     plt.xlabel('Number of Fe atoms in a bin')
#     plt.ylabel('Frequency')
#     plt.title('Distribution of Fe atoms in a cell')
#     plt.savefig('{}_{}'.format(args.output,timestamp))
# except Exception as e:
#     print(e)


# import json
# print(len(return_buckets_coordiate))

# plot_x = []
# plot_y = []
# plot_z = []

# for data in return_buckets_coordiate:
#     plot_x.append(int(data[0]))
#     plot_y.append(int(data[1]))
#     plot_z.append(int(data[2]))

# plot_x = np.array(plot_x)
# plot_y = np.array(plot_y)
# plot_z = np.array(plot_z)

# #imp

# # fig = plt.figure(figsize = (18, 18))
# # seaborn_plot = plt.axes (projection='3d')




# # seaborn_plot.scatter3D (plot_x, plot_y, plot_z)
# # seaborn_plot.set_xlabel ('x')
# # seaborn_plot.set_ylabel ('y')
# # seaborn_plot.set_zlabel ('z')
# # plt.show ()

# #end imp

# #testing
# axes = [int(k)+1, int(k)+1, int(k)+1]
# data_zero = np.zeros(axes, dtype=np.bool)
# for x,y,z in zip(plot_x,plot_y,plot_z):
#     print(x,y,z)
#     data_zero[x-1][y-1][z-1] = True

# data_zero = np.array(data_zero)


# # Control Transparency
# alpha = 0.7
 
# # Control colour
# colors = np.empty(axes + [4], dtype=np.float32)
 
# colors[0] = [1, 1, 1, alpha]  # red
# colors[1] = [1, 1, 1, alpha]   # green
# colors[2] = [1, 1, 1, alpha]   # blue
# colors[3] = [1, 1, 1, alpha]   # yellow
# colors[4] = [1, 1, 1, alpha]  # grey
# colors[5] = [1, 1, 1, alpha] 
# colors[6] = [1, 1, 1, alpha] 
# colors[7] = [1, 1, 1, alpha] 

 
# # Plot figure
# fig = plt.figure(figsize = (18, 18))
# ax = fig.add_subplot(111, projection='3d')
 
# # Voxels is used to customizations of
# # the sizes, positions and colors.
# ax.voxels(data_zero)
# plt.show()


































